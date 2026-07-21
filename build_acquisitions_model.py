#!/usr/bin/env python3
"""Build acquisitions-model.html from the Acquisition Screener Excel.

Reads the 'Forward Looking Model' sheet and regenerates the Acquisitions
page for the dashboard's Forward Model tab. forward-model.html is used as
the design template at build time: its two <tbody> blocks, <title>, <h1>,
and footer source lines are swapped out; all CSS / JS / callout content
carries over untouched so the two pages never drift apart visually.

Spec: docs/superpowers/specs/2026-07-07-acquisitions-ranking-model-design.md

Usage:
    python build_acquisitions_model.py [--xlsx PATH]
"""

import argparse
import datetime
import html
import re
import sys
import warnings
from pathlib import Path

import openpyxl

# read_only workbooks parse lazily, so openpyxl's "extension not supported"
# warnings fire during row iteration, outside any catch_warnings block around
# load_workbook. Silence them process-wide: under weekly-refresh.ps1's
# ErrorActionPreference=Stop, a 2>&1 redirect turns stderr chatter into a
# terminating error.
warnings.filterwarnings("ignore", module="openpyxl")

REPO = Path(__file__).resolve().parent
DEFAULT_XLSX = Path(r"C:\Users\JakeMillenbruck\Subtext\Subtext - Documents"
                    r"\General\Investment\Investment\Research - Analysis"
                    r"\Student Market Analysis\Live Rankings\2026 Rebuild"
                    r"\Acquisition Screener - 2024.xlsx")
TEMPLATE = REPO / "forward-model.html"
OUTPUT = REPO / "acquisitions-model.html"
SHEET = "Forward Looking Model"

PAGE_TITLE = "Forward Looking Model - Acquisition Screener 2024"
PAGE_H1 = "Forward Looking Model - Acquisition Screener"
SOURCE_NAME = "Acquisition Screener &ndash; 2024.xlsx"

# Canonical SharePoint copy of the source workbook. The template carries the
# development screener's live-source link in both footers; swap it for ours.
LIVE_URL = ("https://collegiatedevelopment.sharepoint.com/sites/Subtext/"
            "Shared%20Documents/Forms/AllItems.aspx"
            "?id=%2Fsites%2FSubtext%2FShared%20Documents%2FGeneral"
            "%2FInvestment%2FInvestment%2FResearch%20%2D%20Analysis"
            "%2FStudent%20Market%20Analysis%2FLive%20Rankings"
            "%2F2026%20Rebuild%2FAcquisition%20Screener%20%2D%202024%2Exlsx"
            "&parent=%2Fsites%2FSubtext%2FShared%20Documents%2FGeneral"
            "%2FInvestment%2FInvestment%2FResearch%20%2D%20Analysis"
            "%2FStudent%20Market%20Analysis%2FLive%20Rankings"
            "%2F2026%20Rebuild")
LIVE_LINK = (f'<a class="live-src" href="{LIVE_URL}" target="_blank" '
             f'rel="noopener" style="color:#2c4a8a; font-weight:600;">'
             f'Live source on SharePoint &#8599;</a>')

# 1-based column index -> exact expected header (row 1)
EXPECTED_HEADERS = {
    1: "Current Year Ranking", 2: "Forward Looking Ranking", 3: "Change",
    4: "Market", 5: "Full Time Enrollment", 6: "TTM Prelease",
    7: "PBSH Occupancy", 8: "3 Year Change In Student Bed To Enrollment Ratio",
    9: "TTM Prelease Change", 10: "Three Year Change In Applications",
    11: "POSH Occupancy Last Year", 12: "Growth In FT OoS Undergrads",
    13: "Strongest Variable", 14: "Weakest Variable",
    17: "TTM Prelease", 18: "PBSH Occupancy",
    19: "3 Year Change In Student Bed To Enrollment Ratio",
    20: "TTM Prelease Change", 21: "Three Year Change In Applications",
    22: "POSH Occupancy Last Year", 23: "Growth In FT OoS Undergrads",
    24: "Transactions Last 5", 25: "Transactions Previous 5",
    26: "Construction Last 5", 27: "Construction Previous 5",
    28: "Power 4", 29: "R1", 30: "Rent/Price",
    31: "Current New Property Rent",
}

# Excel variable names (cols M/N) -> short labels used by the tag badges,
# matching the labels the development page uses.
VARIABLE_LABELS = {
    "TTM Prelease": "TTM Prelease",
    "PBSH Occupancy": "PBSH Occ.",
    "3 Year Change In Student Bed To Enrollment Ratio": "3yr Bed/Enroll Δ",
    "TTM Prelease Change": "TTM Prelease Δ",
    "Three Year Change In Applications": "3yr App. Growth",
    "POSH Occupancy Last Year": "POSH Occ. LY",
    "Growth In FT OoS Undergrads": "FT OoS UG Growth",
}


def die(msg):
    sys.exit(f"build_acquisitions_model: ERROR: {msg}")


def is_bad(v):
    """Blank cell or an Excel error string such as #N/A."""
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.startswith("#")


def num(v):
    """Cell as float, or None when blank / error / non-numeric."""
    if is_bad(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_rows(xlsx_path):
    if not xlsx_path.is_file():
        die(f"workbook not found: {xlsx_path}")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        die(f"sheet {SHEET!r} not found; sheets: {wb.sheetnames}")
    ws = wb[SHEET]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        die("sheet is empty")
    for idx, expected in EXPECTED_HEADERS.items():
        got = header[idx - 1] if len(header) >= idx else None
        if got != expected:
            die(f"header mismatch in column {idx}: "
                f"expected {expected!r}, got {got!r}")

    kept, dropped = [], 0
    for row in rows_iter:
        row = (tuple(row) + (None,) * 31)[:31]
        market, fwd_rank = row[3], num(row[1])
        if is_bad(market) or fwd_rank is None:
            if any(not is_bad(v) for v in row):
                dropped += 1
            continue
        kept.append(row)
    if not kept:
        die("no valid data rows after filtering")
    kept.sort(key=lambda r: num(r[1]))
    return kept, dropped


# ---------------------------------------------------------------- formatting

def fmt_int(v):
    return "-" if v is None else f"{v:,.0f}"


def fmt_pct0(v):
    return "-" if v is None else f"{v * 100:.0f}%"


def fmt_pct1(v):
    return "-" if v is None else f"{v * 100:.1f}%"


def fmt_score(v):
    return "-" if v is None else f"{v:.2f}"


def fmt_money(v):
    return "-" if v is None else f"${v:,.0f}"


def market_name(v):
    return html.escape(str(v).strip()).replace(" - ", " – ")


def rank_badge(v, kind):
    if v is None:
        return "<td>-</td>"
    return f'<td><span class="rank rank-{kind}">{v:.0f}</span></td>'


def change_badge(v):
    if v is None:
        return "<td>-</td>"
    n = int(round(v))
    if n > 0:
        return f'<td><span class="change change-up">+{n}</span></td>'
    if n < 0:
        return f'<td><span class="change change-down">{n}</span></td>'
    return '<td><span class="change change-flat">0</span></td>'


def variable_tag(v, kind):
    if is_bad(v):
        return "<td>-</td>"
    label = VARIABLE_LABELS.get(str(v).strip(), html.escape(str(v).strip()))
    return f'<td><span class="tag tag-{kind}">{label}</span></td>'


def metric_td(text, cls):
    return f'<td class="{cls}">{text}</td>' if cls else f"<td>{text}</td>"


def score_td(v):
    if v is None:
        return "<td>-</td>"
    cls = "w-zero" if abs(round(v, 2)) < 0.005 else ("w-pos" if v > 0 else "w-neg")
    return f'<td class="{cls}">{fmt_score(v)}</td>'


def flag_td(v):
    yes = num(v) == 1
    return ('<td><span class="flag-yes">✓</span></td>' if yes
            else '<td><span class="flag-no">–</span></td>')


def make_tercile(values):
    """classify(v) -> metric-lo / metric-mid / metric-hi within one column."""
    nums = sorted(v for v in values if v is not None)
    n = len(nums)
    if n == 0:
        return lambda v: None
    lo_cut, hi_cut = nums[n // 3], nums[(2 * n) // 3]

    def classify(v):
        if v is None:
            return None
        if v < lo_cut:
            return "metric-lo"
        if v >= hi_cut:
            return "metric-hi"
        return "metric-mid"

    return classify


# ------------------------------------------------------------- row rendering

# (column index in row tuple, percent formatter) for the 7 metric columns F-L
METRICS = [(5, fmt_pct0), (6, fmt_pct1), (7, fmt_pct1), (8, fmt_pct1),
           (9, fmt_pct0), (10, fmt_pct0), (11, fmt_pct0)]


def render_values_rows(rows):
    classifiers = {i: make_tercile([num(r[i]) for r in rows]) for i, _ in METRICS}
    out = []
    for r in rows:
        tds = [
            rank_badge(num(r[1]), "forward"),
            rank_badge(num(r[0]), "current"),
            change_badge(num(r[2])),
            f'<td class="left">{market_name(r[3])}</td>',
            f"<td>{fmt_int(num(r[4]))}</td>",
        ]
        for i, fmt in METRICS:
            v = num(r[i])
            tds.append(metric_td(fmt(v), classifiers[i](v)))
        tds.append(variable_tag(r[12], "strong"))
        tds.append(variable_tag(r[13], "weak"))
        out.append("    <tr>" + "".join(tds) + "</tr>")
    return "\n".join(out)


def render_weightings_rows(rows):
    out = []
    for r in rows:
        tds = [
            rank_badge(num(r[1]), "forward"),
            f'<td class="left">{market_name(r[3])}</td>',
        ]
        for i in range(16, 23):                    # Q-W: 7 weighted scores
            tds.append(score_td(num(r[i])))
        for i in range(23, 27):                    # X-AA: 4 counts
            tds.append(f"<td>{fmt_int(num(r[i]))}</td>")
        tds.append(flag_td(r[27]))                 # AB: Power 4
        tds.append(flag_td(r[28]))                 # AC: R1
        tds.append(f"<td>{fmt_pct1(num(r[29]))}</td>")    # AD: Rent/Price
        tds.append(f"<td>{fmt_money(num(r[30]))}</td>")   # AE: New Prop Rent
        out.append("    <tr>" + "".join(tds) + "</tr>")
    return "\n".join(out)


# ------------------------------------------------------------ template swap

def replace_tbody(page, panel_id, rows_html):
    anchor = page.find(f'<div id="{panel_id}"')
    if anchor < 0:
        die(f"template anchor not found: {panel_id}")
    start = page.find("<tbody>", anchor)
    end = page.find("</tbody>", start)
    if start < 0 or end < 0:
        die(f"tbody not found inside {panel_id}")
    return page[:start] + "<tbody>\n" + rows_html + "\n  " + page[end:]


def replace_once(page, pattern, repl, what):
    new, n = re.subn(pattern, repl, page, count=1, flags=re.S)
    if n != 1:
        die(f"template anchor not found: {what}")
    return new


def build(xlsx_path):
    if not TEMPLATE.is_file():
        die(f"template not found: {TEMPLATE}")
    page = TEMPLATE.read_text(encoding="utf-8")
    rows, dropped = load_rows(xlsx_path)

    page = replace_once(page, r"<title>.*?</title>",
                        f"<title>{PAGE_TITLE}</title>", "<title>")
    page = replace_once(page, r"<h1>.*?</h1>",
                        f"<h1>{PAGE_H1}</h1>", "<h1>")
    page = replace_tbody(page, "tab-values", render_values_rows(rows))
    page = replace_tbody(page, "tab-weightings", render_weightings_rows(rows))

    n = page.count("Source: Screener &ndash; 2024.xlsx")
    if n != 2:
        die(f"expected 2 footer source lines, found {n}")
    page = page.replace("Source: Screener &ndash; 2024.xlsx",
                        f"Source: {SOURCE_NAME}")
    today = datetime.date.today().isoformat()
    page, n = re.subn(r"Generated \d{4}-\d{2}-\d{2}", f"Generated {today}", page)
    if n != 2:
        die(f"expected 2 'Generated <date>' stamps, found {n}")

    page, n = re.subn(r'<a class="live-src".*?</a>', LIVE_LINK, page,
                      flags=re.S)
    if n != 2:
        die(f"expected 2 live-source links in template, found {n} "
            "(run build_forward_model.py first)")

    OUTPUT.write_text(page, encoding="utf-8")
    print(f"kept {len(rows)} markets, dropped {dropped} invalid rows")
    print(f"wrote {OUTPUT.name}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                    help="path to Acquisition Screener workbook")
    build(ap.parse_args().xlsx)
