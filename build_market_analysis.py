"""
build_market_analysis.py
------------------------
Data builder for the "Market Analysis - BETA DNU" tab (market-analysis.html).

Replicates the SQL pulls and static reference data of the Excel workbook
  Market Analysis/Market Analysis Template 04092026.xlsm
so the web app can re-render every chart for ANY university instantly,
with no live SQL from the browser.

The workbook's VBA runs per-school queries keyed off Tables!B1 (IPEDS).
This builder runs the same aggregations grouped by IPEDS so one pass
covers every school, then writes market-analysis-data.json.

Sources
  1. StudentResearch Azure SQL (same queries as the workbook macros):
     - YoYPropertyRentGrowth_CH   monthly + yearly rent growth, 4 geo cuts
     - MonthlyPropertyOccupancy_CH monthly + yearly occupancy, 4 geo cuts
     - MonthlyPropertyDataBySF_CH  prelease by year (per report month)
     - MonthlyPropertyData_CH      latest-month property detail (comp table)
     - Properties / School_To_Property  pick-comps property lists
     - PowerFour / Subtext30       benchmark cohorts
  2. Static sheets extracted from the workbook itself:
     - Supply and Demand (FTE, on-campus beds, enrollment by year)
     - RPProperties (pipeline beds: year built, distance, beds)
     - IPEDS Data (applications / admissions / first-time UGs)
     - Manual Enrollment Data
     - Student Affluence lookup block
     - AXIO Crosswalk + University List (search index)

Usage
    python build_market_analysis.py --auth env          (SQLUSER/SQLPASSWORD)
    python build_market_analysis.py --auth integrated   (Entra-joined silent)
    python build_market_analysis.py --auth aad --upn you@subtextliving.com
"""

import argparse
import datetime as dt
import decimal
import getpass
import json
import os
import sys
import time
from pathlib import Path

import pyodbc

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

SERVER = "subtextresearch.database.windows.net"
DATABASE = "StudentResearch"

HERE = Path(__file__).parent
DEFAULT_WORKBOOK_DIR = HERE.parent / "Market Analysis"
OUT_PATH = HERE / "market-analysis-data.json"

YEARLY_MIN_YEAR = 2019   # charts display the last ~6 years
MONTHLY_MIN_YEAR = 2019
COMP_MIN_YEAR = 2020
COMP_PRELEASE_MONTH = 5  # workbook default (Prelease Comp!P2); comp-set only


# ------------------------------------------------------------------ connect

def connect(auth: str, upn: str | None):
    drivers = sorted(
        [d for d in pyodbc.drivers() if d.startswith("ODBC Driver") and "SQL Server" in d],
        reverse=True,
    )
    if not drivers:
        sys.exit("ODBC Driver 17 or 18 for SQL Server is not installed.")
    base = (
        f"Driver={{{drivers[0]}}};"
        f"Server=tcp:{SERVER},1433;Database={DATABASE};"
        "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30;"
    )
    if auth == "integrated":
        cs = base + "Authentication=ActiveDirectoryIntegrated;"
    elif auth == "aad":
        upn = upn or input("Your @subtextliving.com email (Azure AD UPN): ").strip()
        if not upn:
            sys.exit("No UPN provided; aborting.")
        cs = base + f"UID={upn};Authentication=ActiveDirectoryInteractive;"
    elif auth == "sql":
        uid = input("SQL username: ").strip()
        pwd = getpass.getpass("SQL password (not echoed): ")
        if not uid or not pwd:
            sys.exit("Empty username or password; aborting.")
        cs = base + f"UID={uid};PWD={pwd};"
    elif auth == "env":
        uid = os.environ.get("SQLUSER")
        pwd = os.environ.get("SQLPASSWORD")
        if not uid or not pwd:
            sys.exit("--auth env requires SQLUSER and SQLPASSWORD env vars.")
        cs = base + f"UID={uid};PWD={pwd};"
    else:
        sys.exit(f"Unknown --auth mode: {auth}")
    return pyodbc.connect(cs, timeout=60)


def num(v):
    """Decimal/None-safe float with 6dp rounding to keep the JSON small."""
    if v is None:
        return None
    if isinstance(v, decimal.Decimal):
        v = float(v)
    if isinstance(v, float):
        return round(v, 6)
    return v


def run(cn, label, sql):
    t0 = time.time()
    cur = cn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    print(f"  {label}: {len(rows)} rows in {time.time()-t0:.1f}s")
    return rows


# ------------------------------------------------------------------ SQL pulls
# Geo cuts mirror the workbook: all / <1mi / <0.5mi / <1mi & built within 15yrs.
CUTS = {
    "all": "",
    "one_mile": "AND STP.Distance < 1",
    "half_mile": "AND STP.Distance < 0.5",
    "one_mile_new": "AND STP.Distance < 1 AND CAST(P.YearBuilt AS INT)+15 >= year(GetDate())",
}


def pull_monthly(cn, out):
    """Tables macro: monthly YoY rent growth + occupancy, 4 cuts, per school."""
    sql = f"""
    SELECT RG.IPEDS, RG.[Year], RG.[Month],
      SUM(RG.[YoYRent] * RG.[bedsPurposeBuilt]) / NULLIF(SUM(RG.[bedsPurposeBuilt]),0),
      SUM(CASE WHEN STP.Distance < 1 THEN RG.[YoYRent] * RG.[bedsPurposeBuilt] ELSE 0 END) /
        NULLIF(SUM(CASE WHEN STP.Distance < 1 THEN RG.[bedsPurposeBuilt] ELSE 0 END),0),
      SUM(CASE WHEN STP.Distance < .5 THEN RG.[YoYRent] * RG.[bedsPurposeBuilt] ELSE 0 END) /
        NULLIF(SUM(CASE WHEN STP.Distance < .5 THEN RG.[bedsPurposeBuilt] ELSE 0 END),0),
      SUM(CASE WHEN STP.Distance < 1 AND P.YearBuilt+15>=year(GetDate()) THEN RG.[YoYRent] * RG.[bedsPurposeBuilt] ELSE 0 END) /
        NULLIF(SUM(CASE WHEN STP.Distance < 1 AND P.YearBuilt+15>=year(GetDate()) THEN RG.[bedsPurposeBuilt] ELSE 0 END),0)
    FROM StudentResearch.[dbo].[YoYPropertyRentGrowth_CH] RG
    LEFT JOIN StudentResearch.dbo.School_To_Property STP ON RG.PropertyKey = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.Properties P ON RG.PropertyKey = P.[Key]
    WHERE RG.[Year] >= {MONTHLY_MIN_YEAR}
    GROUP BY RG.IPEDS, RG.[Year], RG.[Month]
    ORDER BY RG.IPEDS, RG.[Year], RG.[Month]
    """
    for ipeds, y, m, a, b, c, d in run(cn, "monthly rent growth", sql):
        mk = out.setdefault(str(ipeds), {}).setdefault("monthly_rg", [])
        mk.append([y, m, num(a), num(b), num(c), num(d)])

    sql = f"""
    SELECT MO.IPEDS, MO.[Year], MO.[Month],
      SUM(MO.[Occupancy] * MO.[TotalBeds]) / NULLIF(SUM(MO.[TotalBeds]),0),
      SUM(CASE WHEN STP.Distance < 1 THEN MO.[Occupancy] * MO.[TotalBeds] ELSE 0 END) /
        NULLIF(SUM(CASE WHEN STP.Distance < 1 THEN MO.[TotalBeds] ELSE 0 END),0),
      SUM(CASE WHEN STP.Distance < .5 THEN MO.[Occupancy] * MO.[TotalBeds] ELSE 0 END) /
        NULLIF(SUM(CASE WHEN STP.Distance < .5 THEN MO.[TotalBeds] ELSE 0 END),0),
      SUM(CASE WHEN STP.Distance < 1 AND P.YearBuilt+15>=year(GetDate()) THEN MO.[Occupancy] * MO.[TotalBeds] ELSE 0 END) /
        NULLIF(SUM(CASE WHEN STP.Distance < 1 AND P.YearBuilt+15>=year(GetDate()) THEN MO.[TotalBeds] ELSE 0 END),0)
    FROM StudentResearch.[dbo].[MonthlyPropertyOccupancy_CH] MO
    LEFT JOIN StudentResearch.dbo.School_To_Property STP ON MO.Property_Key = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.Properties P ON MO.Property_Key = P.[Key]
    WHERE MO.[Year] >= {MONTHLY_MIN_YEAR}
    GROUP BY MO.IPEDS, MO.[Year], MO.[Month]
    ORDER BY MO.IPEDS, MO.[Year], MO.[Month]
    """
    for ipeds, y, m, a, b, c, d in run(cn, "monthly occupancy", sql):
        mk = out.setdefault(str(ipeds), {}).setdefault("monthly_occ", [])
        mk.append([y, m, num(a), num(b), num(c), num(d)])


# cut conditions usable inside CASE (no leading AND)
CUT_CASE = {
    "all": "",
    "one_mile": "STP.Distance < 1",
    "half_mile": "STP.Distance < 0.5",
    "one_mile_new": "STP.Distance < 1 AND P.YearBuilt+15 >= year(GetDate())",
}


def _cohort_case(flag, val, wgt, cond):
    cw = f"{flag}" + (f" AND {cond}" if cond else "")
    return (f"SUM(CASE WHEN {cw} THEN {val} * {wgt} ELSE 0 END) / "
            f"NULLIF(SUM(CASE WHEN {cw} THEN {wgt} ELSE 0 END),0)")


def pull_monthly_benchmarks(cn, bench):
    """Monthly cuts for Power 4 + Subtext 30, one table scan per metric:
    8 CASE columns (2 cohorts x 4 cuts) grouped by year+month."""
    cols = []
    for flag in ("PF.Marker = 1", "S.Column1 = 1"):
        for cut in ("all", "one_mile", "half_mile", "one_mile_new"):
            cols.append(_cohort_case(flag, "RG.[YoYRent]", "RG.[bedsPurposeBuilt]", CUT_CASE[cut]))
    sql = f"""
    SELECT RG.[Year], RG.[Month], {", ".join(cols)}
    FROM StudentResearch.[dbo].[YoYPropertyRentGrowth_CH] RG
    LEFT JOIN StudentResearch.dbo.School_To_Property STP ON RG.PropertyKey = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.PowerFour PF ON STP.IPEDS = PF.IPEDS
    LEFT JOIN StudentResearch.dbo.Subtext30 S ON STP.IPEDS = S.IPEDS_ID
    LEFT JOIN StudentResearch.dbo.Properties P ON RG.PropertyKey = P.[Key]
    WHERE RG.[Year] >= {MONTHLY_MIN_YEAR}
    GROUP BY RG.[Year], RG.[Month] ORDER BY RG.[Year], RG.[Month]
    """
    p4rg, s30rg = [], []
    for r in run(cn, "monthly rg bench p4+s30", sql):
        y, m = r[0], r[1]
        p4rg.append([y, m] + [num(v) for v in r[2:6]])
        s30rg.append([y, m] + [num(v) for v in r[6:10]])
    bench.setdefault("p4", {})["monthly_rg"] = p4rg
    bench.setdefault("s30", {})["monthly_rg"] = s30rg

    cols = []
    for flag in ("PF.Marker = 1", "S.Column1 = 1"):
        for cut in ("all", "one_mile", "half_mile", "one_mile_new"):
            cols.append(_cohort_case(flag, "MO.[Occupancy]", "MO.[TotalBeds]", CUT_CASE[cut]))
    sql = f"""
    SELECT MO.[Year], MO.[Month], {", ".join(cols)}
    FROM StudentResearch.[dbo].[MonthlyPropertyOccupancy_CH] MO
    LEFT JOIN StudentResearch.dbo.School_To_Property STP ON MO.Property_Key = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.PowerFour PF ON STP.IPEDS = PF.IPEDS
    LEFT JOIN StudentResearch.dbo.Subtext30 S ON STP.IPEDS = S.IPEDS_ID
    LEFT JOIN StudentResearch.dbo.Properties P ON MO.Property_Key = P.[Key]
    WHERE MO.[Year] >= {MONTHLY_MIN_YEAR}
    GROUP BY MO.[Year], MO.[Month] ORDER BY MO.[Year], MO.[Month]
    """
    p4o, s30o = [], []
    for r in run(cn, "monthly occ bench p4+s30", sql):
        y, m = r[0], r[1]
        p4o.append([y, m] + [num(v) for v in r[2:6]])
        s30o.append([y, m] + [num(v) for v in r[6:10]])
    bench["p4"]["monthly_occ"] = p4o
    bench["s30"]["monthly_occ"] = s30o


def _cut_case(cut_cond, val, wgt):
    if not cut_cond:
        return f"SUM({val} * {wgt}) / NULLIF(SUM({wgt}),0)"
    return (f"SUM(CASE WHEN {cut_cond} THEN {val} * {wgt} ELSE 0 END) / "
            f"NULLIF(SUM(CASE WHEN {cut_cond} THEN {wgt} ELSE 0 END),0)")


CUT_ORDER = ["all", "one_mile", "half_mile", "one_mile_new"]


def pull_yearly(cn, out, bench):
    """Rent Growth / Occ / Prelease Comp. By Year macros. One table scan per
    metric for the schools (4 cut CASE columns, grouped by STP.IPEDS + year)
    and one per metric for both benchmarks (8 columns, grouped by year)."""
    # ---- schools
    cols = [_cut_case(CUT_CASE[c], "CAST(RG.[YoYRent] AS FLOAT)", "RG.[bedsPurposeBuilt]") for c in CUT_ORDER]
    sql = f"""
    SELECT STP.IPEDS, RG.[Year], {", ".join(cols)}
    FROM StudentResearch.[dbo].[YoYPropertyRentGrowth_CH] RG
    JOIN StudentResearch.dbo.School_To_Property STP ON RG.PropertyKey = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.Properties P ON RG.PropertyKey = P.[Key]
    WHERE RG.[Year] >= {YEARLY_MIN_YEAR}
    GROUP BY STP.IPEDS, RG.[Year]
    """
    for r in run(cn, "yearly rg schools", sql):
        blk = out.setdefault(str(r[0]), {}).setdefault("rg_yearly", {})
        for i, c in enumerate(CUT_ORDER):
            if r[2 + i] is not None:
                blk.setdefault(c, {})[str(r[1])] = num(r[2 + i])

    cols = [_cut_case(CUT_CASE[c], "MO.[Occupancy]", "MO.[TotalBeds]") for c in CUT_ORDER]
    sql = f"""
    SELECT STP.IPEDS, MO.[Year], {", ".join(cols)}
    FROM StudentResearch.[dbo].[MonthlyPropertyOccupancy_CH] MO
    JOIN StudentResearch.dbo.School_To_Property STP ON MO.Property_Key = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.Properties P ON MO.Property_Key = P.[Key]
    WHERE MO.[Year] >= {YEARLY_MIN_YEAR}
    GROUP BY STP.IPEDS, MO.[Year]
    """
    for r in run(cn, "yearly occ schools", sql):
        blk = out.setdefault(str(r[0]), {}).setdefault("occ_yearly", {})
        for i, c in enumerate(CUT_ORDER):
            if r[2 + i] is not None:
                blk.setdefault(c, {})[str(r[1])] = num(r[2 + i])

    # prelease keeps the month dimension so the UI month selector works;
    # workbook formula is 1 - SUM(num)/SUM(denom)
    def pre_col(cond):
        if not cond:
            return "SUM(CAST(D.[PreleaseNum] AS FLOAT)) / NULLIF(SUM(D.[PreleaseDenom]),0)"
        return (f"SUM(CASE WHEN {cond} THEN CAST(D.[PreleaseNum] AS FLOAT) ELSE 0 END) / "
                f"NULLIF(SUM(CASE WHEN {cond} THEN D.[PreleaseDenom] ELSE 0 END),0)")
    cols = [pre_col(CUT_CASE[c]) for c in CUT_ORDER]
    sql = f"""
    SELECT STP.IPEDS, year(D.[MonthDate]), month(D.[MonthDate]), {", ".join(cols)}
    FROM StudentResearch.[dbo].[MonthlyPropertyDataBySF_CH] D
    JOIN StudentResearch.dbo.School_To_Property STP ON D.Property_Key = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.Properties P ON D.Property_Key = P.[Key]
    WHERE D.PreleaseDenom > 0 AND year(D.[MonthDate]) >= {YEARLY_MIN_YEAR}
    GROUP BY STP.IPEDS, year(D.[MonthDate]), month(D.[MonthDate])
    """
    for r in run(cn, "yearly prelease schools", sql):
        blk = out.setdefault(str(r[0]), {}).setdefault("pre_yearly", {})
        for i, c in enumerate(CUT_ORDER):
            if r[3 + i] is not None:
                blk.setdefault(c, {})[f"{r[1]}-{r[2]}"] = num(1 - r[3 + i])

    # ---- benchmarks (both cohorts x 4 cuts per scan)
    def cohort_cols(val, wgt):
        cols = []
        for flag in ("PF.Marker = 1", "S.Column1 = 1"):
            for c in CUT_ORDER:
                cond = flag + (f" AND {CUT_CASE[c]}" if CUT_CASE[c] else "")
                cols.append(f"SUM(CASE WHEN {cond} THEN {val} * {wgt} ELSE 0 END) / "
                            f"NULLIF(SUM(CASE WHEN {cond} THEN {wgt} ELSE 0 END),0)")
        return cols

    sql = f"""
    SELECT RG.[Year], {", ".join(cohort_cols("CAST(RG.[YoYRent] AS FLOAT)", "RG.[bedsPurposeBuilt]"))}
    FROM StudentResearch.[dbo].[YoYPropertyRentGrowth_CH] RG
    JOIN StudentResearch.dbo.School_To_Property STP ON RG.PropertyKey = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.PowerFour PF ON STP.IPEDS = PF.IPEDS
    LEFT JOIN StudentResearch.dbo.Subtext30 S ON STP.IPEDS = S.IPEDS_ID
    LEFT JOIN StudentResearch.dbo.Properties P ON RG.PropertyKey = P.[Key]
    WHERE RG.[Year] >= {YEARLY_MIN_YEAR}
    GROUP BY RG.[Year]
    """
    for r in run(cn, "yearly rg bench", sql):
        for j, key in enumerate(("p4", "s30")):
            blk = bench.setdefault(key, {}).setdefault("rg_yearly", {})
            for i, c in enumerate(CUT_ORDER):
                v = r[1 + j * 4 + i]
                if v is not None:
                    blk.setdefault(c, {})[str(r[0])] = num(v)

    sql = f"""
    SELECT MO.[Year], {", ".join(cohort_cols("MO.[Occupancy]", "MO.[TotalBeds]"))}
    FROM StudentResearch.[dbo].[MonthlyPropertyOccupancy_CH] MO
    JOIN StudentResearch.dbo.School_To_Property STP ON MO.Property_Key = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.PowerFour PF ON STP.IPEDS = PF.IPEDS
    LEFT JOIN StudentResearch.dbo.Subtext30 S ON STP.IPEDS = S.IPEDS_ID
    LEFT JOIN StudentResearch.dbo.Properties P ON MO.Property_Key = P.[Key]
    WHERE MO.[Year] >= {YEARLY_MIN_YEAR}
    GROUP BY MO.[Year]
    """
    for r in run(cn, "yearly occ bench", sql):
        for j, key in enumerate(("p4", "s30")):
            blk = bench.setdefault(key, {}).setdefault("occ_yearly", {})
            for i, c in enumerate(CUT_ORDER):
                v = r[1 + j * 4 + i]
                if v is not None:
                    blk.setdefault(c, {})[str(r[0])] = num(v)

    def cohort_pre_cols():
        cols = []
        for flag in ("PF.Marker = 1", "S.Column1 = 1"):
            for c in CUT_ORDER:
                cond = flag + (f" AND {CUT_CASE[c]}" if CUT_CASE[c] else "")
                cols.append(f"SUM(CASE WHEN {cond} THEN CAST(D.[PreleaseNum] AS FLOAT) ELSE 0 END) / "
                            f"NULLIF(SUM(CASE WHEN {cond} THEN D.[PreleaseDenom] ELSE 0 END),0)")
        return cols

    sql = f"""
    SELECT year(D.[MonthDate]), month(D.[MonthDate]), {", ".join(cohort_pre_cols())}
    FROM StudentResearch.[dbo].[MonthlyPropertyDataBySF_CH] D
    JOIN StudentResearch.dbo.School_To_Property STP ON D.Property_Key = STP.PropertyKey
    LEFT JOIN StudentResearch.dbo.PowerFour PF ON STP.IPEDS = PF.IPEDS
    LEFT JOIN StudentResearch.dbo.Subtext30 S ON STP.IPEDS = S.IPEDS_ID
    LEFT JOIN StudentResearch.dbo.Properties P ON D.Property_Key = P.[Key]
    WHERE D.PreleaseDenom > 0 AND year(D.[MonthDate]) >= {YEARLY_MIN_YEAR}
    GROUP BY year(D.[MonthDate]), month(D.[MonthDate])
    """
    for r in run(cn, "yearly pre bench", sql):
        for j, key in enumerate(("p4", "s30")):
            blk = bench.setdefault(key, {}).setdefault("pre_yearly", {})
            for i, c in enumerate(CUT_ORDER):
                v = r[2 + j * 4 + i]
                if v is not None:
                    blk.setdefault(c, {})[f"{r[0]}-{r[1]}"] = num(1 - v)


def pull_properties(cn, out):
    """Pick Comps! macro: candidate comp properties per school."""
    sql = """
    SELECT STP.IPEDS, P.[key], CAST(P.[name] AS NVARCHAR(255)),
           P.[YearBuilt], P.[bedsPurposeBuilt], P.[milesToClosestCampus]
    FROM StudentResearch.dbo.Properties P
    JOIN StudentResearch.dbo.School_To_Property STP ON P.[key] = STP.PropertyKey
    """
    for ipeds, key, name, built, beds, dist in run(cn, "pick-comps properties", sql):
        lst = out.setdefault(str(ipeds), {}).setdefault("properties", [])
        lst.append({
            "key": key, "name": name, "built": built,
            "beds": beds, "dist": num(dist),
        })


def pull_comp_series(cn, out):
    """Per-property-per-year numerator/denominator pairs so the client can
    aggregate any user-picked comp set exactly like the workbook does."""
    sql = f"""
    SELECT STP.IPEDS, RG.PropertyKey, RG.[Year],
           SUM(CAST(RG.[YoYRent] AS FLOAT) * RG.[bedsPurposeBuilt]), SUM(RG.[bedsPurposeBuilt])
    FROM StudentResearch.[dbo].[YoYPropertyRentGrowth_CH] RG
    JOIN StudentResearch.dbo.School_To_Property STP ON RG.PropertyKey = STP.PropertyKey
    WHERE RG.[Year] >= {COMP_MIN_YEAR}
    GROUP BY STP.IPEDS, RG.PropertyKey, RG.[Year]
    """
    for ipeds, pk, y, n, d in run(cn, "comp rg num/den", sql):
        c = out.setdefault(str(ipeds), {}).setdefault("comp_rg", {}).setdefault(str(pk), {})
        c[str(y)] = [num(n), num(d)]
    sql = f"""
    SELECT STP.IPEDS, MO.Property_Key, MO.[Year],
           SUM(MO.[Occupancy] * MO.[TotalBeds]), SUM(MO.[TotalBeds])
    FROM StudentResearch.[dbo].[MonthlyPropertyOccupancy_CH] MO
    JOIN StudentResearch.dbo.School_To_Property STP ON MO.Property_Key = STP.PropertyKey
    WHERE MO.[Year] >= {COMP_MIN_YEAR}
    GROUP BY STP.IPEDS, MO.Property_Key, MO.[Year]
    """
    for ipeds, pk, y, n, d in run(cn, "comp occ num/den", sql):
        c = out.setdefault(str(ipeds), {}).setdefault("comp_occ", {}).setdefault(str(pk), {})
        c[str(y)] = [num(n), num(d)]
    sql = f"""
    SELECT STP.IPEDS, D.Property_Key, year(D.[MonthDate]),
           SUM(CAST(D.[PreleaseNum] AS FLOAT)), SUM(D.[PreleaseDenom])
    FROM StudentResearch.[dbo].[MonthlyPropertyDataBySF_CH] D
    JOIN StudentResearch.dbo.School_To_Property STP ON D.Property_Key = STP.PropertyKey
    WHERE D.PreleaseDenom > 0 AND month(D.[MonthDate]) = {COMP_PRELEASE_MONTH}
      AND year(D.[MonthDate]) >= {COMP_MIN_YEAR}
    GROUP BY STP.IPEDS, D.Property_Key, year(D.[MonthDate])
    """
    for ipeds, pk, y, n, d in run(cn, "comp prelease num/den", sql):
        c = out.setdefault(str(ipeds), {}).setdefault("comp_pre", {}).setdefault(str(pk), {})
        c[str(y)] = [num(n), num(d)]


def pull_latest_detail(cn):
    """Competitive Set macro: last calendar month per-property detail."""
    sql = """
    SELECT D.Property_Key,
           P.[units], P.[bedsPurposeBuilt], P.[YearBuilt],
           CASE WHEN D.[RateDenom]=0 THEN NULL ELSE D.[RateNum]/D.[RateDenom] END,
           RG.[YoYRent],
           CASE WHEN D.[VacDenom]=0 THEN NULL ELSE D.[OccNum]/CAST(D.[VacDenom] AS FLOAT) END,
           CASE WHEN D.[PreleaseDenom]=0 THEN NULL ELSE D.[PreleaseNum]/CAST(D.[PreleaseDenom] AS FLOAT) END
    FROM StudentResearch.dbo.MonthlyPropertyData_CH D
    LEFT JOIN StudentResearch.dbo.Properties P ON D.Property_key = P.[key]
    LEFT JOIN StudentResearch.dbo.YoYPropertyRentGrowth_CH RG
      ON D.Property_key = RG.PropertyKey
     AND month(D.MonthDate) = RG.[Month] AND year(D.MonthDate) = RG.[Year]
    WHERE datediff(mm, D.MonthDate, getdate()) = 1
    """
    latest = {}
    for pk, units, beds, built, rent, rg, occ, pre in run(cn, "latest property detail", sql):
        latest[str(pk)] = {
            "units": units, "beds": beds, "built": built,
            "rent_bed": num(rent), "rg": num(rg), "occ": num(occ), "pre": num(pre),
        }
    return latest


# ------------------------------------------------------------- workbook side

def find_workbook(path_arg):
    if path_arg:
        return Path(path_arg)
    cands = sorted(DEFAULT_WORKBOOK_DIR.glob("*.xlsm"), key=lambda p: p.stat().st_mtime)
    if not cands:
        sys.exit(f"No .xlsm found in {DEFAULT_WORKBOOK_DIR}")
    return cands[-1]


def _clean(v):
    """Excel cached values: turn #N/A-style strings and 'NULL' into None."""
    if isinstance(v, str):
        s = v.strip()
        if s in ("#N/A", "#VALUE!", "#REF!", "#DIV/0!", "NULL", "NA", ""):
            return None
        try:
            f = float(s)
            return int(f) if f.is_integer() else f
        except ValueError:
            return s
    if isinstance(v, dt.datetime):
        return v.year
    if isinstance(v, decimal.Decimal):
        return float(v)
    if isinstance(v, float):
        return round(v, 6)
    return v


def extract_workbook(wb_path):
    import openpyxl
    print(f"Extracting static sheets from {wb_path.name} ...")
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # --- AXIO Crosswalk: university name -> IPEDS
    name_to_ipeds = {}
    for r in wb["AXIO Crosswalk"].iter_rows(min_row=2, max_col=4, values_only=True):
        nm, _t, _uid, ip = (_clean(x) for x in r)
        if nm and ip:
            name_to_ipeds[nm] = int(ip)

    # --- University List: the searchable market set
    uni_list = []
    for (nm,) in wb["University List"].iter_rows(min_row=2, max_col=1, values_only=True):
        nm = _clean(nm)
        if nm and nm in name_to_ipeds:
            uni_list.append({"name": nm, "ipeds": name_to_ipeds[nm]})

    # --- Supply and Demand: per-IPEDS yearly (FTE real, on-campus beds, etc.)
    # Columns: C=University, G=Total Existing Beds, H=New Supply,
    #          I=Total Fall Enrollment, L=Undergrad, M=Grad,
    #          AI=University Housing Existing Beds (on-campus),
    #          AV=FTE Real, AW=Year, AX=IPEDS, AY=Subtext30, AZ=Power4
    sd = {}
    sd_bench = {"s30": {}, "p4": {}}
    for r in wb["Supply and Demand"].iter_rows(min_row=5, max_col=52, values_only=True):
        name = _clean(r[2])
        year = _clean(r[48])
        ipeds = _clean(r[49])
        if year is None:
            continue
        year = int(year)
        fte = _clean(r[47])
        onc = _clean(r[34])
        row = {
            "fte": fte, "on_campus": onc,
            "enroll": _clean(r[8]), "beds_total": _clean(r[6]),
            "new_supply": _clean(r[7]), "undergrad": _clean(r[11]), "grad": _clean(r[12]),
        }
        if ipeds is not None:
            sd.setdefault(str(int(ipeds)), {})[str(year)] = row
        s30 = _clean(r[50]) == 1
        p4 = _clean(r[51]) == 1
        for flag, key in ((s30, "s30"), (p4, "p4")):
            if flag:
                agg = sd_bench[key].setdefault(str(year), {"fte": 0, "on_campus": 0})
                if isinstance(fte, (int, float)):
                    agg["fte"] += fte
                if isinstance(onc, (int, float)):
                    agg["on_campus"] += onc

    # --- RPProperties: pipeline/PBSH beds. AA=YearBuilt, AE=Beds,
    #     AJ=DistToCampus, AN=IPEDS, AO=Subtext30, AP=Power4
    rp = {}
    rp_bench = {"s30": [], "p4": []}
    for r in wb["RPProperties"].iter_rows(min_row=2, max_col=44, values_only=True):
        built = _clean(r[26])
        beds = _clean(r[30])
        dist = _clean(r[35])
        ipeds = _clean(r[39])
        if built is None or beds is None:
            continue
        rec = [int(built), beds, dist]
        if ipeds is not None:
            rp.setdefault(str(int(ipeds)), []).append(rec)
        if _clean(r[40]) == 1:
            rp_bench["s30"].append(rec)
        if _clean(r[41]) == 1:
            rp_bench["p4"].append(rec)

    # --- IPEDS Data: applications/admissions/first-time UGs (+ benchmarks)
    # C=UnitID, D=Year, E=FTE, F=ADMSSN, G=APPLCN, H=FirstTimeUGrads,
    # I=Subtext30 flag, J=Power4 flag, K=Count (7 = complete panel)
    wanted = {u["ipeds"] for u in uni_list}
    ipeds_hist = {}
    ipeds_bench = {"s30": {}, "p4": {}}
    for r in wb["IPEDS Data"].iter_rows(min_row=2, max_col=11, values_only=True):
        uid = _clean(r[2]); yr = _clean(r[3])
        if uid is None or yr is None or yr < 2015:
            continue
        fte, adm, app, ftug = (_clean(r[i]) for i in (4, 5, 6, 7))
        s30 = _clean(r[8]) == 1
        p4 = _clean(r[9]) == 1 and _clean(r[10]) == 7
        if int(uid) in wanted:
            ipeds_hist.setdefault(str(int(uid)), {})[str(int(yr))] = [fte, adm, app, ftug]
        for flag, key in ((s30, "s30"), (p4, "p4")):
            if flag:
                agg = ipeds_bench[key].setdefault(str(int(yr)), [0, 0, 0, 0])
                for i, v in enumerate((fte, adm, app, ftug)):
                    if isinstance(v, (int, float)):
                        agg[i] += v

    # --- Manual Enrollment Data: A=IPEDS, B=Year, E=FT Enrollment,
    #     I=S30, J=P4, K=Count, M=Yes Or No
    man = {}
    man_rows = []
    max_count = 0
    for r in wb["Manual Enrollment Data"].iter_rows(min_row=2, max_col=13, values_only=True):
        ip = _clean(r[0]); yr = _clean(r[1]); fte = _clean(r[4])
        if ip is None or yr is None:
            continue
        cnt = _clean(r[10]) or 0
        max_count = max(max_count, cnt if isinstance(cnt, (int, float)) else 0)
        man_rows.append((int(ip), int(yr), fte, _clean(r[8]) == 1,
                         _clean(r[9]) == 1, cnt, _clean(r[12]) == 1))
    man_bench = {"s30": {}, "p4": {}}
    for ip, yr, fte, s30, p4, cnt, yes in man_rows:
        man.setdefault(str(ip), {})[str(yr)] = fte
        if s30 and cnt == max_count and isinstance(fte, (int, float)):
            man_bench["s30"][str(yr)] = man_bench["s30"].get(str(yr), 0) + fte
        if p4 and yes and isinstance(fte, (int, float)):
            man_bench["p4"][str(yr)] = man_bench["p4"].get(str(yr), 0) + fte

    # --- Student Affluence lookup block (manual comp set in the template)
    affl = []
    for r in wb["Student Affluence"].iter_rows(min_row=4, max_row=18, min_col=34,
                                               max_col=48, values_only=True):
        name = _clean(r[3])
        if not name:
            continue
        affl.append({
            "name": name, "ipeds": _clean(r[4]),
            "fte": _clean(r[5]), "fte_5yr_prior": _clean(r[6]), "cagr": _clean(r[7]),
            "tuition": _clean(r[8]), "rent": _clean(r[9]), "income": _clean(r[10]),
            "tr_share": _clean(r[11]), "hi_income_students": _clean(r[13]),
            "hi_income_beds": _clean(r[14]),
        })

    wb.close()
    print(f"  crosswalk={len(name_to_ipeds)} unis={len(uni_list)} sd={len(sd)} "
          f"rp={sum(len(v) for v in rp.values())} ipeds_hist={len(ipeds_hist)} "
          f"manual={len(man)} affluence={len(affl)}")
    return {
        "universities": uni_list,
        "supply_demand": sd,
        "supply_demand_bench": sd_bench,
        "rp": rp,
        "rp_bench_s30": rp_bench["s30"],
        "rp_bench_p4": rp_bench["p4"],
        "ipeds_history": ipeds_hist,
        "ipeds_bench": ipeds_bench,
        "manual_enrollment": man,
        "manual_bench": man_bench,
        "affluence": affl,
    }


# ------------------------------------------------------------------ main

def main():
    ap = argparse.ArgumentParser(description="Build market-analysis-data.json")
    ap.add_argument("--auth", choices=["aad", "sql", "env", "integrated"], default="env")
    ap.add_argument("--upn", default=None)
    ap.add_argument("--workbook", default=None,
                    help="Path to the Market Analysis .xlsm (default: newest in ../Market Analysis)")
    ap.add_argument("--skip-sql", action="store_true",
                    help="Only refresh the workbook-derived static sections.")
    args = ap.parse_args()

    wb_path = find_workbook(args.workbook)
    static = extract_workbook(wb_path)

    markets = {}
    bench = {}
    latest = {}
    if not args.skip_sql:
        print(f"Connecting to {SERVER} / {DATABASE} ...")
        cn = connect(args.auth, args.upn)
        print("Pulling SQL aggregates (this is the slow part) ...")
        pull_monthly(cn, markets)
        pull_monthly_benchmarks(cn, bench)
        pull_yearly(cn, markets, bench)
        pull_properties(cn, markets)
        pull_comp_series(cn, markets)
        latest = pull_latest_detail(cn)
        cn.close()
    elif OUT_PATH.exists():
        prev = json.loads(OUT_PATH.read_text(encoding="utf-8"))
        markets = prev.get("markets", {})
        bench = prev.get("benchmarks", {})
        latest = prev.get("latest_detail", {})
        print("  (--skip-sql: reusing SQL sections from the existing JSON)")

    payload = {
        "generated": dt.datetime.now().isoformat(timespec="seconds"),
        "workbook": wb_path.name,
        "comp_prelease_month": COMP_PRELEASE_MONTH,
        "markets": markets,
        "benchmarks": bench,
        "latest_detail": latest,
        **static,
    }
    OUT_PATH.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    size_mb = OUT_PATH.stat().st_size / 1e6
    print(f"Wrote {OUT_PATH.name} ({size_mb:.1f} MB), "
          f"{len(markets)} markets with SQL series, {len(static['universities'])} searchable.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
