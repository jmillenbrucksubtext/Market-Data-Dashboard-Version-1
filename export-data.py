"""
export-data.py
--------------
Connects to StudentResearch, runs the 9 dashboard queries, and writes a
single data.json that the static site consumes.

Run modes:
  Local - Azure AD interactive (PREFERRED, no password in shell history):
      python export-data.py            # defaults to --auth aad
      python export-data.py --auth aad

  Local - SQL login, prompted via getpass (not echoed, not in argv):
      python export-data.py --auth sql

  CI / scripted - env vars (GitHub Actions cron, see .github/workflows/refresh.yml):
      set SQLUSER=...   set SQLPASSWORD=...
      python export-data.py --auth env

Inputs:
  --auth aad : prompts once for your @subtextliving.com email, then browser sign-in.
  --auth sql : prompts for SQL username + password (getpass, never echoed).
  --auth env : reads SQLUSER + SQLPASSWORD env vars.
  ODBC Driver 18 for SQL Server installed.

Output:
  ./data.json  - one document with metadata + a key per metric table.
"""

from __future__ import annotations

import argparse
import datetime as dt
import decimal
import getpass
import json
import os
import re
import sys
import threading
from pathlib import Path

import pyodbc

SERVER = "subtextresearch.database.windows.net"
DATABASE = "StudentResearch"

OUTPUT = Path(__file__).parent / "data.json"

# Same query bodies as the views in d.data-model/metric-views.sql and
# the M expressions in f.powerbi/build-walkthrough.md. Single source of
# truth for the metric definitions.
QUERIES: dict[str, str] = {
    "penetration": """
        SELECT
            m.[Key]                                  AS market_key,
            mr.beds_purpose_built                    AS total_beds,
            mr.enr_total                             AS total_enrollment,
            CASE WHEN mr.enr_total > 0
                 THEN CAST(mr.beds_purpose_built AS DECIMAL(18,6)) / mr.enr_total
                 ELSE NULL END                       AS penetration_ratio,
            CASE
                WHEN mr.enr_total IS NULL OR mr.enr_total = 0 THEN 'N/A'
                WHEN CAST(mr.beds_purpose_built AS DECIMAL(18,6)) / mr.enr_total < 0.30 THEN 'Under-supplied'
                WHEN CAST(mr.beds_purpose_built AS DECIMAL(18,6)) / mr.enr_total > 0.55 THEN 'Over-supplied'
                ELSE 'Balanced'
            END                                      AS market_band,
            mr.snapshot_date                         AS data_as_of
        FROM dbo.Markets m
        JOIN (
            SELECT market_key, MAX(snapshot_date) AS max_snap
            FROM dbo.MarketReports
            WHERE market_key IS NOT NULL
            GROUP BY market_key
        ) latest ON latest.market_key = m.[Key]
        JOIN dbo.MarketReports mr
               ON  mr.market_key    = latest.market_key
               AND mr.snapshot_date = latest.max_snap
    """,
    # prelease_velocity: last 3 cycles only, and only the top 50 markets by
    # tracked-bed volume. The multi-year line chart is unreadable beyond
    # ~30 series anyway; users will filter to a market in the UI.
    "prelease_velocity": """
        WITH cycle_calc AS (
            SELECT
                mr.market_key,
                mr.snapshot_date,
                mr.prelease,
                mr.beds_tracked_by_prelease,
                CASE WHEN MONTH(mr.snapshot_date) <= 8
                     THEN YEAR(mr.snapshot_date)
                     ELSE YEAR(mr.snapshot_date) + 1 END AS leasing_cycle,
                DATEPART(ISO_WEEK, mr.snapshot_date) AS week_of_cycle
            FROM dbo.MarketReports mr
            WHERE mr.prelease IS NOT NULL
              AND mr.beds_tracked_by_prelease > 0
              AND mr.market_key IS NOT NULL
        ),
        max_cycle AS (SELECT MAX(leasing_cycle) AS yr FROM cycle_calc),
        top_markets AS (
            SELECT TOP 50 market_key
            FROM cycle_calc, max_cycle
            WHERE leasing_cycle = max_cycle.yr
            GROUP BY market_key
            ORDER BY SUM(beds_tracked_by_prelease) DESC
        )
        SELECT
            cc.market_key,
            cc.leasing_cycle,
            cc.week_of_cycle,
            AVG(CAST(cc.prelease AS DECIMAL(18,6)))  AS prelease_pct,
            SUM(cc.beds_tracked_by_prelease)         AS beds_tracked,
            MAX(cc.snapshot_date)                    AS data_as_of
        FROM cycle_calc cc, max_cycle
        WHERE cc.leasing_cycle >= max_cycle.yr - 2
          AND cc.market_key IN (SELECT market_key FROM top_markets)
        GROUP BY cc.market_key, cc.leasing_cycle, cc.week_of_cycle
    """,
    "existing_beds": """
        SELECT
            m.[Key]                  AS market_key,
            mr.beds_purpose_built    AS existing_beds,
            mr.snapshot_date         AS data_as_of
        FROM dbo.Markets m
        JOIN (
            SELECT market_key, MAX(snapshot_date) AS max_snap
            FROM dbo.MarketReports
            WHERE market_key IS NOT NULL
            GROUP BY market_key
        ) latest ON latest.market_key = m.[Key]
        JOIN dbo.MarketReports mr
               ON  mr.market_key    = latest.market_key
               AND mr.snapshot_date = latest.max_snap
    """,
    "pipeline_beds": """
        SELECT
            m.[Key]                          AS market_key,
            mr.beds_lease_up                 AS beds_lease_up,
            mr.beds_under_construction       AS beds_under_construction,
            mr.beds_planned                  AS beds_planned,
            mr.beds_pipeline                 AS beds_pipeline_total,
            mr.snapshot_date                 AS data_as_of
        FROM dbo.Markets m
        JOIN (
            SELECT market_key, MAX(snapshot_date) AS max_snap
            FROM dbo.MarketReports
            WHERE market_key IS NOT NULL
            GROUP BY market_key
        ) latest ON latest.market_key = m.[Key]
        JOIN dbo.MarketReports mr
               ON  mr.market_key    = latest.market_key
               AND mr.snapshot_date = latest.max_snap
    """,
    # enrollment_total: only schools mapped to a market we're tracking. Drops
    # the ~3,200 schools with no marketKey. Enrollment value/year sourced
    # from the latest Enrollments_Manual row per school (much fresher than
    # the snapshot in Schools_Denormal, which was sometimes ~2014).
    "enrollment_total": """
        SELECT
            sd.schoolKey                                                 AS university_key,
            sd.name                                                      AS university_name,
            sd.marketKey                                                 AS market_key,
            COALESCE(em.Total_Enrollment, sd.enrollmentTotal)            AS total_enrollment,
            sd.enrollmentFullTimeUndergraduate
              + sd.enrollmentPartTimeUndergraduate                       AS total_undergrad,
            COALESCE(em.Year,             sd.enrollmentYear)             AS academic_year
        FROM dbo.Schools_Denormal sd
        JOIN dbo.Markets m ON m.[Key] = sd.marketKey
        LEFT JOIN dbo.IPEDS_CH_Crosswalk cx
               ON cx.[Key]     = sd.schoolKey
              AND cx.marketKey = sd.marketKey
        OUTER APPLY (
            SELECT TOP 1 Total_Enrollment, Year
            FROM dbo.Enrollments_Manual
            WHERE IPEDS_ID = cx.IPEDs
              AND Total_Enrollment > 0
            ORDER BY Year DESC
        ) em
        WHERE COALESCE(em.Total_Enrollment, sd.enrollmentTotal) > 0
    """,
    # Full multi-year enrollment history per university (Total / FTE /
    # Freshman / etc.) - drives the Enrollment History charts on the
    # Market tab. Mirrors load_enrollment_history.py. Same dedupe pattern
    # as enrollment_trend: collapse Enrollments_Manual to one row per
    # (IPEDS, Year) and dedupe the crosswalk join.
    "enrollment_history": """
        WITH e AS (
            SELECT IPEDS_ID,
                   MIN(University)                AS university_name,
                   Year,
                   MAX(Total_Enrollment)          AS total_enrollment,
                   MAX(Full_Time_Enrollment)      AS full_time_enrollment,
                   MAX(Undergrad_Enrollment)      AS undergrad_enrollment,
                   MAX(Graduate_Enrollment)       AS graduate_enrollment,
                   MAX(Freshman_Enrollment)       AS freshman_enrollment
            FROM dbo.Enrollments_Manual
            WHERE Total_Enrollment > 0
            GROUP BY IPEDS_ID, Year
        ),
        cx AS (SELECT DISTINCT IPEDs, marketKey FROM dbo.IPEDS_CH_Crosswalk)
        SELECT
            e.IPEDS_ID                AS ipeds_id,
            e.university_name,
            cx.marketKey              AS market_key,
            e.Year                    AS year_,
            e.total_enrollment,
            e.full_time_enrollment,
            e.undergrad_enrollment,
            e.graduate_enrollment,
            e.freshman_enrollment
        FROM e
        LEFT JOIN cx ON cx.IPEDs = e.IPEDS_ID
        WHERE e.Year >= 2015
        ORDER BY e.IPEDS_ID, e.Year
    """,
    # university_info: per-school institutional stats from Schools_Denormal -
    # on-campus beds, tuition, admissions funnel, housing rates, student
    # profile. Drives the University Information tab on the market page.
    # Schools_Denormal is school x year; keep the latest year with real
    # enrollment, one row per (marketKey, schoolKey). Mirrors
    # load_university_info.py. Zeros in rate / income columns mean "not
    # reported" and are treated as missing by the UI.
    "university_info": """
        WITH ranked AS (
            SELECT sd.*,
                   ROW_NUMBER() OVER (
                       PARTITION BY sd.marketKey, sd.schoolKey
                       ORDER BY sd.enrollmentYear DESC
                   ) AS rn
            FROM dbo.Schools_Denormal sd
            WHERE sd.marketKey IS NOT NULL
              AND COALESCE(sd.enrollmentTotal, 0) > 0
        )
        SELECT
            sd.marketKey                                   AS market_key,
            sd.schoolKey                                   AS school_key,
            sd.name                                        AS university_name,
            sd.shortName                                   AS short_name,
            sd.isPublic                                    AS is_public,
            sd.enrollmentYear                              AS enrollment_year,
            sd.enrollmentTotal                             AS enrollment_total,
            sd.enrollmentFullTimeUndergraduate             AS enr_ft_undergrad,
            sd.enrollmentPartTimeUndergraduate             AS enr_pt_undergrad,
            sd.enrollmentFullTimeGraduate                  AS enr_ft_grad,
            sd.enrollmentPartTimeGraduate                  AS enr_pt_grad,
            sd.enrollmentUndergradInStatePct               AS pct_in_state,
            sd.enrollmentUndergradOutOfStatePct            AS pct_out_of_state,
            sd.enrollmentUndergradOnCampusPct              AS pct_on_campus,
            sd.enrollmentUndergradOffCampusPct             AS pct_off_campus,
            sd.bedsOnCampusReported                        AS beds_on_campus_reported,
            sd.bedsOnCampusComputed                        AS beds_on_campus_computed,
            sd.rateOnCampusRoomYearly                      AS rate_room_yearly,
            sd.rateOnCampusBoardYearly                     AS rate_board_yearly,
            sd.rateOnCampusRoomMonthlyAvg                  AS rate_room_monthly,
            sd.costHousingDelta                            AS cost_housing_delta,
            sd.tuitionInState                              AS tuition_in_state,
            sd.tuitionOutOfState                           AS tuition_out_of_state,
            sd.costPerCreditHourUndergradInState           AS credit_hour_in_state,
            sd.costPerCreditHourUndergradOutOfState        AS credit_hour_out_of_state,
            sd.appliedFirstTimeFirstYear                   AS applied_first_year,
            sd.admittedFirstTimeFirstYear                  AS admitted_first_year,
            sd.admittedFirstTimeFirstYearPct               AS admit_rate,
            sd.enrolledFirstTimeFirstYear                  AS enrolled_first_year,
            sd.appliedTransfer                             AS applied_transfer,
            sd.admittedTransfer                            AS admitted_transfer,
            sd.admittedTransferPct                         AS transfer_admit_rate,
            sd.enrolledTransfer                            AS enrolled_transfer,
            sd.studentAgeAvg                               AS student_age_avg,
            sd.amountParentIncomeAvg                       AS parent_income_avg,
            sd.amountParentIncomeMed                       AS parent_income_med,
            sd.hasDormsCoed                                AS has_dorms_coed,
            sd.hasDormsMen                                 AS has_dorms_men,
            sd.hasDormsWomen                               AS has_dorms_women,
            sd.hasApartmentsMarried                        AS has_apts_married,
            sd.hasApartmentsSingle                         AS has_apts_single,
            sd.hasHousingDisabled                          AS has_housing_disabled,
            sd.hasHousingIntl                              AS has_housing_intl,
            sd.hasHousingGreek                             AS has_housing_greek,
            sd.hasHousingCoop                              AS has_housing_coop
        FROM ranked sd
        WHERE sd.rn = 1
        ORDER BY sd.marketKey, sd.name
    """,
    "enrollment_trend": """
        -- Per-university 1yr and 5yr enrollment trends. Dedupes at two
        -- spots that historically fanned out:
        --   1. Enrollments_Manual sometimes has multiple rows for the same
        --      (IPEDS_ID, Year) - different totals. We collapse with MAX so
        --      one row per (IPEDS, Year).
        --   2. IPEDS_CH_Crosswalk occasionally has multiple rows for the
        --      same IPEDS → market mapping. We dedupe via DISTINCT in the
        --      crosswalk subquery so the LEFT JOIN never doubles output.
        WITH e AS (
            SELECT IPEDS_ID,
                   MIN(University)        AS University,
                   Year,
                   MAX(Total_Enrollment)  AS Total_Enrollment
            FROM dbo.Enrollments_Manual
            WHERE Total_Enrollment > 0
            GROUP BY IPEDS_ID, Year
        ),
        latest AS (SELECT IPEDS_ID, MAX(Year) AS yr_max FROM e GROUP BY IPEDS_ID),
        cx_unique AS (
            SELECT DISTINCT IPEDs, marketKey FROM dbo.IPEDS_CH_Crosswalk
        )
        SELECT
            cur.IPEDS_ID                                AS ipeds_id,
            cur.University                              AS university_name,
            cx.marketKey                                AS market_key,
            cur.Year                                    AS current_year,
            cur.Total_Enrollment                        AS current_enrollment,
            prev1.Total_Enrollment                      AS prev_year_enrollment,
            CASE WHEN prev1.Total_Enrollment > 0
                 THEN CAST(cur.Total_Enrollment AS DECIMAL(18,6)) / prev1.Total_Enrollment - 1
                 ELSE NULL END                          AS yoy_change,
            CASE WHEN prev5.Total_Enrollment > 0
                 THEN POWER(CAST(cur.Total_Enrollment AS DECIMAL(18,8)) / prev5.Total_Enrollment, 1.0/5.0) - 1
                 ELSE NULL END                          AS cagr_5yr
        FROM e cur
        JOIN latest l            ON l.IPEDS_ID = cur.IPEDS_ID AND l.yr_max = cur.Year
        LEFT JOIN e prev1        ON prev1.IPEDS_ID = cur.IPEDS_ID AND prev1.Year = cur.Year - 1
        LEFT JOIN e prev5        ON prev5.IPEDS_ID = cur.IPEDS_ID AND prev5.Year = cur.Year - 5
        LEFT JOIN cx_unique cx   ON cx.IPEDs = cur.IPEDS_ID
        LEFT JOIN dbo.Markets m  ON m.[Key]  = cx.marketKey
    """,
    # fte_history: per-market full-time enrollment at three points in time -
    # current snapshot, ~1 year prior, and a 2021-22 academic-year baseline.
    # Drives the FTE YoY KPI tile and the two FTE qualifiers
    # ("FTE growth YoY positive", "FTE growth since 2022 above 3%").
    "fte_history": """
        WITH current_snap AS (
            SELECT
                latest.market_key,
                latest.max_snap          AS current_date_,
                mr.enr_full_time         AS current_fte
            FROM (
                SELECT market_key, MAX(snapshot_date) AS max_snap
                FROM dbo.MarketReports
                WHERE market_key IS NOT NULL AND enr_full_time IS NOT NULL
                GROUP BY market_key
            ) latest
            JOIN dbo.MarketReports mr
                   ON  mr.market_key    = latest.market_key
                  AND mr.snapshot_date = latest.max_snap
        ),
        -- Anchor "prior" to the most recent snapshot whose enr_full_time
        -- DIFFERS from current. This sidesteps IPEDS publish cadence: if a
        -- school's FTE didn't change in the last 12 months, the weekly
        -- MarketReports rows all carry the same value forward and a literal
        -- T-1yr lookup would always return current_fte. Compare-to-last-change
        -- always finds a real prior cohort.
        prior_yr AS (
            SELECT
                cs.market_key,
                mr.snapshot_date         AS prior_yr_date,
                mr.enr_full_time         AS prior_yr_fte,
                ROW_NUMBER() OVER (
                    PARTITION BY cs.market_key
                    ORDER BY mr.snapshot_date DESC
                ) AS rn
            FROM current_snap cs
            JOIN dbo.MarketReports mr
                   ON  mr.market_key     = cs.market_key
                  AND mr.enr_full_time  IS NOT NULL
                  AND mr.enr_full_time  <> cs.current_fte
                  AND mr.snapshot_date  <  cs.current_date_
        ),
        y2022 AS (
            SELECT
                cs.market_key,
                mr.snapshot_date         AS y2022_date,
                mr.enr_full_time         AS y2022_fte,
                ROW_NUMBER() OVER (
                    PARTITION BY cs.market_key
                    ORDER BY ABS(DATEDIFF(DAY, '2022-01-01', mr.snapshot_date))
                ) AS rn
            FROM current_snap cs
            JOIN dbo.MarketReports mr
                   ON  mr.market_key     = cs.market_key
                  AND mr.enr_full_time  IS NOT NULL
                  AND mr.snapshot_date BETWEEN '2021-07-01' AND '2022-12-31'
        )
        SELECT
            cs.market_key,
            cs.current_date_                 AS current_snapshot,
            cs.current_fte,
            py.prior_yr_date                 AS prior_year_snapshot,
            py.prior_yr_fte                  AS prior_year_fte,
            CASE WHEN py.prior_yr_fte > 0
                 THEN CAST(cs.current_fte AS DECIMAL(18,6)) / py.prior_yr_fte - 1
                 ELSE NULL END               AS yoy_fte_growth,
            y22.y2022_date                   AS baseline_2022_snapshot,
            y22.y2022_fte                    AS baseline_2022_fte,
            CASE WHEN y22.y2022_fte > 0
                 THEN CAST(cs.current_fte AS DECIMAL(18,6)) / y22.y2022_fte - 1
                 ELSE NULL END               AS fte_growth_since_2022,
            cs.current_date_                 AS data_as_of
        FROM current_snap cs
        LEFT JOIN prior_yr py
               ON py.market_key = cs.market_key AND py.rn = 1
        LEFT JOIN y2022 y22
               ON y22.market_key = cs.market_key AND y22.rn = 1
    """,
    "avg_rent": """
        SELECT
            m.[Key]                          AS market_key,
            mr.rate_avg                      AS avg_rent_per_bed,
            mr.beds_tracked_by_occupancy     AS bed_weight,
            mr.snapshot_date                 AS data_as_of
        FROM dbo.Markets m
        JOIN (
            SELECT market_key, MAX(snapshot_date) AS max_snap
            FROM dbo.MarketReports
            WHERE market_key IS NOT NULL
            GROUP BY market_key
        ) latest ON latest.market_key = m.[Key]
        JOIN dbo.MarketReports mr
               ON  mr.market_key    = latest.market_key
               AND mr.snapshot_date = latest.max_snap
        WHERE mr.rate_avg IS NOT NULL
    """,
    "rent_yoy": """
        WITH current_snap AS (
            SELECT
                latest.market_key,
                latest.max_snap AS current_date_,
                mr.rate_avg     AS current_avg_rent
            FROM (
                SELECT market_key, MAX(snapshot_date) AS max_snap
                FROM dbo.MarketReports
                WHERE market_key IS NOT NULL AND rate_avg IS NOT NULL
                GROUP BY market_key
            ) latest
            JOIN dbo.MarketReports mr
                   ON  mr.market_key    = latest.market_key
                   AND mr.snapshot_date = latest.max_snap
        ),
        prior AS (
            SELECT
                cs.market_key,
                cs.current_date_,
                cs.current_avg_rent,
                prior_mr.snapshot_date AS prior_date_,
                prior_mr.rate_avg      AS prior_avg_rent,
                ROW_NUMBER() OVER (
                    PARTITION BY cs.market_key
                    ORDER BY ABS(DATEDIFF(DAY, DATEADD(YEAR, -1, cs.current_date_), prior_mr.snapshot_date))
                ) AS rn
            FROM current_snap cs
            JOIN dbo.MarketReports prior_mr
                   ON  prior_mr.market_key = cs.market_key
                  AND prior_mr.rate_avg   IS NOT NULL
                  AND prior_mr.snapshot_date BETWEEN DATEADD(DAY, -395, cs.current_date_)
                                                 AND DATEADD(DAY, -335, cs.current_date_)
        )
        SELECT
            p.market_key,
            p.current_date_           AS current_snapshot,
            p.prior_date_             AS prior_snapshot,
            p.current_avg_rent,
            p.prior_avg_rent,
            CASE WHEN p.prior_avg_rent > 0
                 THEN CAST(p.current_avg_rent AS DECIMAL(18,6)) / p.prior_avg_rent - 1
                 ELSE NULL END        AS yoy_rent_growth,
            p.current_date_           AS data_as_of
        FROM prior p
        JOIN dbo.Markets m ON m.[Key] = p.market_key
        WHERE p.rn = 1
    """,
    # properties: every purpose-built property mapped to a market. Drives the
    # property list and the map on market.html. Drop isShadow=1 (shadow market)
    # and rows without lat/long since the map can't render them.
    "properties": """
        WITH plan_latest AS (
            -- Latest snapshot per property in PlanReports
            SELECT property_key, MAX(snapshot_date) AS max_snap
            FROM dbo.PlanReports
            WHERE property_key IS NOT NULL AND prelease IS NOT NULL
            GROUP BY property_key
        ),
        plan_agg AS (
            -- Bed-weighted prelease from plan-level data at the latest snapshot.
            -- PlanReports.prelease coverage is ~6.6M rows vs Properties.prelease
            -- which is null for ~99.7% of rows - this is the actual source.
            SELECT
                pr.property_key,
                SUM(CAST(pr.prelease AS DECIMAL(18,6)) * pr.beds_purpose_built)
                    / NULLIF(SUM(pr.beds_purpose_built), 0) AS weighted_prelease,
                MAX(pr.snapshot_date) AS prelease_as_of
            FROM dbo.PlanReports pr
            JOIN plan_latest pl
                   ON  pl.property_key = pr.property_key
                  AND pl.max_snap      = pr.snapshot_date
            WHERE pr.prelease IS NOT NULL
              AND pr.beds_purpose_built > 0
            GROUP BY pr.property_key
        )
        SELECT
            p.[Key]                                 AS property_key,
            p.marketKey                             AS market_key,
            p.name                                  AS property_name,
            p.type                                  AS property_type,
            p.phase                                 AS phase,
            p.street1,
            p.city,
            p.state,
            p.latitude,
            p.longitude,
            p.bedsPurposeBuilt                      AS beds,
            p.units,
            p.yearBuilt,
            p.occupancy,
            COALESCE(p.prelease, pa.weighted_prelease) AS prelease,
            pa.prelease_as_of                       AS prelease_as_of,
            p.rateAvg                               AS avg_rent,
            p.ratePerSfAvg                          AS avg_rent_per_sf,
            p.concessionsValue,
            p.hasConcessions,
            p.milesToClosestCampus,
            p.currentGoogleReviewAvg,
            p.currentGoogleReviews                  AS google_review_count,
            p.phone,
            p.webSiteUrl                            AS website,
            CAST(p.lastReportDate AS DATETIME2)     AS last_report_date
        FROM dbo.Properties p
        LEFT JOIN plan_agg pa ON pa.property_key = p.[Key]
        WHERE p.marketKey IS NOT NULL
          AND COALESCE(p.isShadow, 0) = 0
    """,
    # plans: every floor plan tied to a purpose-built property we already
    # track. Drives the property.html floor-plan table.
    "plans": """
        SELECT
            pl.[Key]                     AS plan_key,
            pl.propertyKey               AS property_key,
            pl.marketKey                 AS market_key,
            pl.name                      AS plan_name,
            pl.unitTypeName              AS unit_type_name,
            pl.format,
            pl.isStudio                  AS is_studio,
            pl.bedrooms,
            pl.bathrooms,
            pl.areaSf                    AS area_sf,
            pl.bedsPurposeBuilt          AS beds_in_plan,
            pl.occupancy,
            pl.prelease,
            pl.rate,
            pl.ratePerSf                 AS rate_per_sf,
            pl.hasConcessions            AS has_concessions,
            pl.concessionsValue          AS concessions_value,
            pl.concessionsNotes          AS concessions_notes,
            CAST(pl.reportDate AS DATETIME2) AS report_date
        FROM dbo.Plans pl
        JOIN dbo.Properties p ON p.[Key] = pl.propertyKey
        WHERE p.marketKey IS NOT NULL
          AND COALESCE(p.isShadow, 0) = 0
    """,
    # campus_locations: every school with lat/long that maps to a market we
    # track. The market_detail page uses these to drop campus pins on the map.
    # Enrollment value/year now comes from the latest Enrollments_Manual row
    # per school (typically 2023–2024), falling back to Schools_Denormal's
    # stale snapshot only when no IPEDS history exists.
    "campus_locations": """
        SELECT
            sd.marketKey                                            AS market_key,
            sd.schoolKey                                            AS school_key,
            cx.IPEDs                                                AS ipeds_id,
            sd.name                                                 AS university_name,
            sd.latitude                                             AS campus_lat,
            sd.longitude                                            AS campus_lng,
            COALESCE(em.Total_Enrollment, sd.enrollmentTotal)       AS total_enrollment,
            COALESCE(em.Year,             sd.enrollmentYear)        AS enrollment_year
        FROM dbo.Schools_Denormal sd
        JOIN dbo.Markets m ON m.[Key] = sd.marketKey
        LEFT JOIN dbo.IPEDS_CH_Crosswalk cx
               ON cx.[Key]     = sd.schoolKey
              AND cx.marketKey = sd.marketKey
        OUTER APPLY (
            SELECT TOP 1 Total_Enrollment, Year
            FROM dbo.Enrollments_Manual
            WHERE IPEDS_ID = cx.IPEDs
              AND Total_Enrollment > 0
            ORDER BY Year DESC
        ) em
        WHERE sd.latitude  IS NOT NULL
          AND sd.longitude IS NOT NULL
    """,
    # scorecard is the master per-market table. anchor_university = largest
    # university by enrollment in that market (fallback "City, ST" if no school
    # is mapped). is_subtext30 = 1 if any university in this market is one of
    # the Subtext 30 focus universities.
    #
    # total_enrollment / enr_full_time come from the CURRENT academic year in
    # Enrollments_Manual (summed across the market's mapped universities,
    # each at its own latest reported year), NOT from MarketReports - the
    # snapshot columns there carry last year's IPEDS cohort forward (per data
    # science, 2026-06). MarketReports remains the fallback for markets with
    # no Enrollments_Manual coverage.
    "scorecard": """
        WITH latest AS (
            SELECT market_key, MAX(snapshot_date) AS max_snap
            FROM dbo.MarketReports
            WHERE market_key IS NOT NULL
            GROUP BY market_key
        ),
        anchor AS (
            SELECT m.[Key] AS market_key, a.name AS anchor_university
            FROM dbo.Markets m
            OUTER APPLY (
                SELECT TOP 1 sd.name
                FROM dbo.Schools_Denormal sd
                WHERE sd.marketKey = m.[Key] AND sd.enrollmentTotal > 0
                ORDER BY sd.enrollmentTotal DESC, sd.name
            ) a
        ),
        s30 AS (
            SELECT DISTINCT cx.marketKey AS market_key
            FROM dbo.Subtext30 s
            JOIN dbo.IPEDS_CH_Crosswalk cx ON cx.IPEDs = s.IPEDS_ID
            WHERE cx.marketKey IS NOT NULL
            UNION
            -- Name-match fallback for Subtext30 rows whose IPEDS_ID has no
            -- crosswalk entry. Penn State has two IPEDS IDs (Subtext30 lists
            -- 495767; the crosswalk + Enrollments_Manual use 214777), so its
            -- ID join misses. Known-gaps.md gap #4.
            SELECT DISTINCT sd.marketKey
            FROM dbo.Subtext30 s
            JOIN dbo.Schools_Denormal sd ON sd.name = s.University
            WHERE sd.marketKey IS NOT NULL
              AND NOT EXISTS (
                  SELECT 1 FROM dbo.IPEDS_CH_Crosswalk cx
                  WHERE cx.IPEDs = s.IPEDS_ID AND cx.marketKey IS NOT NULL
              )
        ),
        -- Current-year enrollment per market from Enrollments_Manual: each
        -- university contributes its latest reported year, per column (total
        -- and full-time can publish on different cadences).
        cxu AS (SELECT DISTINCT IPEDs, marketKey FROM dbo.IPEDS_CH_Crosswalk),
        em_tv AS (
            SELECT e.IPEDS_ID, MAX(e.Total_Enrollment) AS total_enr
            FROM dbo.Enrollments_Manual e
            JOIN (
                SELECT IPEDS_ID, MAX(Year) AS yr
                FROM dbo.Enrollments_Manual
                WHERE Total_Enrollment > 0
                GROUP BY IPEDS_ID
            ) ly ON ly.IPEDS_ID = e.IPEDS_ID AND e.Year = ly.yr
            GROUP BY e.IPEDS_ID
        ),
        em_fv AS (
            SELECT e.IPEDS_ID, MAX(e.Full_Time_Enrollment) AS fte
            FROM dbo.Enrollments_Manual e
            JOIN (
                SELECT IPEDS_ID, MAX(Year) AS yr
                FROM dbo.Enrollments_Manual
                WHERE Full_Time_Enrollment > 0
                GROUP BY IPEDS_ID
            ) ly ON ly.IPEDS_ID = e.IPEDS_ID AND e.Year = ly.yr
            GROUP BY e.IPEDS_ID
        ),
        em_mkt AS (
            SELECT cxu.marketKey         AS market_key,
                   SUM(em_tv.total_enr)  AS em_total,
                   SUM(em_fv.fte)        AS em_fte
            FROM cxu
            LEFT JOIN em_tv ON em_tv.IPEDS_ID = cxu.IPEDs
            LEFT JOIN em_fv ON em_fv.IPEDS_ID = cxu.IPEDs
            GROUP BY cxu.marketKey
        )
        SELECT
            m.[Key]                                              AS market_key,
            COALESCE(a.anchor_university,
                     m.city + ', ' + m.stateAbbr)                AS anchor_university,
            m.city                                               AS city,
            m.stateAbbr                                          AS state_abbr,
            m.region                                             AS region,   -- regional grouping (Midwest, South, ...)
            CASE WHEN s30.market_key IS NOT NULL THEN 1 ELSE 0 END
                                                                 AS is_subtext30,
            CASE WHEN mr.enr_total > 0
                 THEN CAST(mr.beds_purpose_built AS DECIMAL(18,6)) / mr.enr_total END
                                                                 AS penetration_ratio,
            mr.beds_purpose_built                                AS existing_beds,
            mr.beds_lease_up,
            mr.beds_under_construction,
            mr.beds_planned,
            mr.beds_pipeline                                     AS beds_pipeline_total,
            COALESCE(NULLIF(em.em_total, 0), mr.enr_total)       AS total_enrollment,
            COALESCE(NULLIF(em.em_fte,   0), mr.enr_full_time)   AS enr_full_time,
            mr.rate_avg                                          AS avg_rent_per_bed,
            mr.occupancy                                         AS occupancy,
            mr.prelease                                          AS prelease,
            mr.snapshot_date                                     AS data_as_of
        FROM dbo.Markets m
        JOIN latest               ON latest.market_key = m.[Key]
        JOIN dbo.MarketReports mr ON mr.market_key = latest.market_key
                                  AND mr.snapshot_date = latest.max_snap
        LEFT JOIN anchor a        ON a.market_key   = m.[Key]
        LEFT JOIN s30             ON s30.market_key = m.[Key]
        LEFT JOIN em_mkt em       ON em.market_key  = m.[Key]
    """,
    # market_history: one anchor snapshot per (market, year) from 2020+.
    # Each year is anchored to the SAME calendar date as the latest
    # available snapshot, so the year-over-year comparison is apples to
    # apples (e.g., 2026-05-09 vs 2025-05-09 vs 2024-05-09 ...). This
    # moves automatically as the weekly refresh advances the latest
    # snapshot. Feeds the deck-style multi-year charts on market.html.
    "market_history": (
        "WITH la AS (SELECT MONTH(MAX(snapshot_date)) AS mo, DAY(MAX(snapshot_date)) AS dy FROM dbo.MarketReports), "
        "a AS ("
        "SELECT mr.market_key, YEAR(mr.snapshot_date) AS yr, mr.snapshot_date, mr.rate_avg, mr.occupancy, mr.prelease, "
        "mr.beds_purpose_built, mr.beds_lease_up, mr.beds_under_construction, mr.beds_planned, mr.beds_pipeline, "
        "ROW_NUMBER() OVER (PARTITION BY mr.market_key, YEAR(mr.snapshot_date) "
        "ORDER BY ABS(DATEDIFF(DAY, mr.snapshot_date, "
        "DATEFROMPARTS(YEAR(mr.snapshot_date), la.mo, "
        "CASE WHEN la.dy > DAY(EOMONTH(DATEFROMPARTS(YEAR(mr.snapshot_date), la.mo, 1))) "
        "THEN DAY(EOMONTH(DATEFROMPARTS(YEAR(mr.snapshot_date), la.mo, 1))) ELSE la.dy END)))) AS rn "
        "FROM dbo.MarketReports mr CROSS JOIN la "
        "WHERE mr.market_key IS NOT NULL AND YEAR(mr.snapshot_date) >= 2020"
        ") "
        "SELECT market_key, yr AS year_, snapshot_date AS data_as_of, rate_avg AS avg_rent_per_bed, "
        "occupancy, prelease, beds_purpose_built AS existing_beds, beds_lease_up, beds_under_construction, "
        "beds_planned, beds_pipeline AS beds_pipeline_total "
        "FROM a WHERE rn = 1 ORDER BY market_key, yr"
    ),
    # property_history: per-property annual anchored snapshot pulled from
    # dbo.PlanReports. Bed-weighted across plans within each property for
    # rate, rate_per_sf, occupancy, prelease. Drives the comp-set
    # time-series charts on market.html. See load_property_history.py for
    # the standalone equivalent.
    "property_history": """
        WITH la AS (
            SELECT MONTH(MAX(snapshot_date)) AS mo, DAY(MAX(snapshot_date)) AS dy
            FROM dbo.PlanReports WHERE property_key IS NOT NULL
        ),
        snaps AS (
            SELECT DISTINCT pr.property_key, pr.snapshot_date,
                            YEAR(pr.snapshot_date) AS yr
            FROM dbo.PlanReports pr
            WHERE pr.property_key IS NOT NULL AND YEAR(pr.snapshot_date) >= 2020
        ),
        anchored AS (
            SELECT s.property_key, s.snapshot_date, s.yr,
                ROW_NUMBER() OVER (
                    PARTITION BY s.property_key, s.yr
                    ORDER BY ABS(DATEDIFF(DAY, s.snapshot_date,
                                DATEFROMPARTS(s.yr, la.mo,
                                  CASE WHEN la.dy > DAY(EOMONTH(DATEFROMPARTS(s.yr, la.mo, 1)))
                                       THEN DAY(EOMONTH(DATEFROMPARTS(s.yr, la.mo, 1)))
                                       ELSE la.dy END)))
                ) AS rn
            FROM snaps s CROSS JOIN la
        ),
        chosen AS (
            SELECT property_key, yr, snapshot_date FROM anchored WHERE rn = 1
        )
        SELECT
            pr.property_key,
            c.yr AS year_,
            c.snapshot_date AS data_as_of,
            CASE WHEN SUM(pr.beds_purpose_built) > 0
                 THEN SUM(pr.rate * pr.beds_purpose_built)
                      / NULLIF(SUM(pr.beds_purpose_built), 0)
            END AS avg_rent_per_bed,
            CASE WHEN SUM(pr.beds_purpose_built) > 0
                 THEN SUM(pr.rate_per_sf * pr.beds_purpose_built)
                      / NULLIF(SUM(pr.beds_purpose_built), 0)
            END AS avg_rent_per_sf,
            CASE WHEN SUM(pr.beds_purpose_built) > 0
                 THEN SUM(pr.occupancy * pr.beds_purpose_built)
                      / NULLIF(SUM(pr.beds_purpose_built), 0)
            END AS occupancy,
            CASE WHEN SUM(pr.beds_purpose_built) > 0
                 THEN SUM(pr.prelease * pr.beds_purpose_built)
                      / NULLIF(SUM(pr.beds_purpose_built), 0)
            END AS prelease,
            SUM(pr.beds_purpose_built) AS beds
        FROM chosen c
        JOIN dbo.PlanReports pr
               ON pr.property_key = c.property_key AND pr.snapshot_date = c.snapshot_date
        GROUP BY pr.property_key, c.yr, c.snapshot_date
        ORDER BY pr.property_key, c.yr
    """,
}


def to_jsonable(value):
    """Convert pyodbc row values to JSON-serializable primitives."""
    if value is None:
        return None
    if isinstance(value, decimal.Decimal):
        return float(value)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return value


# ============================================================
# Pursuit markets - Subtext's active "Markets - Pursuing" pipeline
# ============================================================
# Source: dbo.ProjectCosts, project_stage = 'Markets - Pursuing', at the latest
# UpdateDate snapshot (the CRM re-snapshots ~weekly). Each row is a market-level
# pursuit titled "<University> - <City>[ - P3]". We map each to a dashboard
# market_key via a city (+state) crosswalk built from the scorecard, then stamp
# is_pursuit / pursuit_deals onto the scorecard so the Industry-page toggle can
# filter the map + list. Auto-updates each refresh as the pipeline changes.
PURSUIT_STAGE = "Markets - Pursuing"
# University keyword -> state, to disambiguate same-named cities
# (the pipeline has both "Columbia, MO" (Missouri) and "Columbia, SC" (S. Carolina)).
_PURSUIT_UNI_STATE = {"missouri": "MO", "south carolina": "SC"}
# City aliases where the CRM's city label differs from dbo.Markets.city.
# Storrs is the village UConn sits in; the dashboard files it under Mansfield, CT.
_PURSUIT_CITY_ALIAS = {"storrs": "mansfield"}


def _parse_pursuit_title(title: str) -> tuple[str, str]:
    """From a '<University> - <City>[ - P3]' pursuit title, return (university, city)."""
    t = re.sub(r"\s*-\s*P3\s*$", "", (title or "").strip(), flags=re.I)
    if " - " in t:
        uni, city = t.rsplit(" - ", 1)
        return uni.strip(), city.strip()
    return t, t


def compute_pursuit_markets(cur, scorecard: list[dict]) -> dict[int, int]:
    """Return {market_key: pursuit_deal_count} for markets in 'Markets - Pursuing'.

    Non-university entries (e.g. 'Dallas / Fort Worth') and any title that
    doesn't resolve to a tracked market are logged and skipped - non-fatal, so a
    pipeline hiccup never blocks the weekly refresh."""
    by_city_state: dict[tuple[str, str], int] = {}
    by_city: dict[str, int] = {}
    for r in scorecard:
        c = (r.get("city") or "").strip().lower()
        st = (r.get("state_abbr") or "").strip().upper()
        if not c:
            continue
        by_city_state[(c, st)] = r["market_key"]
        by_city.setdefault(c, r["market_key"])

    try:
        cur.execute(
            "SELECT MAX(UpdateDate) FROM dbo.ProjectCosts WHERE project_stage = ?",
            PURSUIT_STAGE)
        max_date = cur.fetchone()[0]
        if max_date is None:
            print("  pursuit_markets: no 'Markets - Pursuing' rows (skipping)")
            return {}
        cur.execute(
            "SELECT title FROM dbo.ProjectCosts "
            "WHERE project_stage = ? AND UpdateDate = ?", PURSUIT_STAGE, max_date)
        titles = [r[0] for r in cur.fetchall() if r[0]]
    except pyodbc.Error as e:
        print(f"  pursuit_markets SKIPPED (query failed: {e})")
        return {}

    counts: dict[int, int] = {}
    unmatched: list[str] = []
    for title in titles:
        uni, city = _parse_pursuit_title(title)
        key = _PURSUIT_CITY_ALIAS.get(city.lower(), city.lower())
        st = next((s for kw, s in _PURSUIT_UNI_STATE.items() if kw in uni.lower()), None)
        mk = by_city_state.get((key, st)) if st else None
        if mk is None:
            mk = by_city.get(key)
        if mk is None:
            unmatched.append(title)
            continue
        counts[mk] = counts.get(mk, 0) + 1

    snap = str(max_date)[:10]
    print(f"  pursuit_markets: {len(titles)} '{PURSUIT_STAGE}' rows "
          f"({snap}) -> {len(counts)} dashboard markets")
    if unmatched:
        # Surface new/unmapped pursuits so an alias can be added next refresh.
        print(f"    {len(unmatched)} unmatched (no tracked market): "
              + "; ".join(unmatched))
    return counts


# ============================================================
# Active markets - hardcoded 3-stage CRM view (Upcoming / Assessing / Pursuing)
# ============================================================
# TEMPORARY: pinned by market_key to match the CRM's "Markets - *" board exactly
# (captured 2026-07-01). This supersedes the SQL-derived pursuit flag on the
# Industry map because the ProjectCosts snapshot lags the CRM by ~a snapshot and
# lumps the parallel "- P3" track in with the student-housing (SH) pipeline.
# compute_pursuit_markets() above is left in place as the basis for a future
# live 3-stage sync; until then these lists are the source of truth.
#   TODO: replace with a live feed off dbo.ProjectCosts project_stage, filtering
#   out '- P3' titles and mapping Upcoming/Assessing/Pursuing per stage.
ACTIVE_MARKET_STATUS: dict[int, str] = {
    # Markets - Pursuing (18)
    155: "pursuing",   # Ohio State - Columbus
    960: "pursuing",   # Oregon State - Corvallis
    2375: "pursuing",  # Penn State - State College
    2417: "pursuing",  # Rutgers - New Brunswick
    2418: "pursuing",  # UConn - Storrs/Mansfield
    2400: "pursuing",  # Virginia Tech - Blacksburg
    2344: "pursuing",  # UNC Chapel Hill
    86: "pursuing",    # Missouri - Columbia
    1804: "pursuing",  # LSU - Baton Rouge
    27: "pursuing",    # UT Austin
    1012: "pursuing",  # Utah - Salt Lake City
    2294: "pursuing",  # Maryland - College Park
    977: "pursuing",   # Colorado Boulder
    567: "pursuing",   # Kentucky - Lexington
    732: "pursuing",   # Syracuse
    118: "pursuing",   # Indiana - Bloomington
    12: "pursuing",    # Georgia - Athens
    11: "pursuing",    # Clemson
    # Markets - Assessing (1)
    334: "assessing",  # Wisconsin - Madison
    # Markets - Upcoming (5)
    975: "upcoming",   # San Diego (SDSU)
    940: "upcoming",   # Washington - Seattle
    15: "upcoming",    # Ole Miss - Oxford
    383: "upcoming",   # Kansas - Lawrence
    418: "upcoming",   # South Florida - Tampa
}


# ============================================================
# Forward Looking Model - rank per market (from forward-model.html)
# ============================================================
# The data-science team drops a self-contained forward-model.html into this
# folder (embedded on the Industry page as the "Forward Model" view). We parse
# its market ranking and stamp fwd_rank onto the scorecard so the Industry table
# can show each market's screener rank. Re-read every refresh, so a re-dropped
# HTML updates the column automatically.
FORWARD_HTML = Path(__file__).parent / "forward-model.html"
# Screener names that are shorter/variant than the dashboard's anchor name, or
# ambiguous across multiple campuses - map to the intended flagship. Values are
# canonical anchor names (resolved through the scorecard, so no hard-coded keys).
_FORWARD_ALIAS = {
    "pennsylvania state university": "Penn State",
    "university at buffalo state university of new york": "University at Buffalo SUNY",
    "university of illinois urbana champaign": "University of Illinois at Urbana-Champaign",
    "university of north carolina": "University of North Carolina at Chapel Hill",
    "university of massachusetts": "University of Massachusetts Amherst",
    "indiana university": "Indiana University Bloomington",
    "university of minnesota": "University of Minnesota Twin Cities",
}


def _norm_name(s: str) -> str:
    s = (s or "").lower().replace("–", "-").replace("—", "-")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def compute_forward_ranks(scorecard: list[dict]) -> dict[int, int]:
    """Return {market_key: forward_rank} parsed from forward-model.html.

    Non-fatal: a missing or unreadable file just yields no ranks."""
    if not FORWARD_HTML.exists():
        print("  forward_ranks: forward-model.html not found (skipping)")
        return {}
    try:
        html = FORWARD_HTML.read_text(encoding="utf-8")
    except OSError as e:
        print(f"  forward_ranks SKIPPED (read failed: {e})")
        return {}
    # Read the Values tab only, so each market is counted once.
    start, end = html.find('id="tab-values"'), html.find('id="tab-weightings"')
    section = html[start:end] if start != -1 and end != -1 else html

    anchor_to_key: dict[str, int] = {}
    city_to_key: dict[str, int] = {}
    for r in scorecard:
        anchor_to_key[_norm_name(r["anchor_university"])] = r["market_key"]
        city_to_key.setdefault(_norm_name(r["city"]), r["market_key"])

    def resolve(name: str):
        n = _norm_name(name)
        if n in _FORWARD_ALIAS:                       # variant / flagship lock
            n = _norm_name(_FORWARD_ALIAS[n])
        if n in anchor_to_key:                        # exact anchor
            return anchor_to_key[n]
        parts = re.split(r"\s+-\s+", name.replace("–", "-"), maxsplit=1)
        uni = _norm_name(parts[0])
        city = _norm_name(parts[1]) if len(parts) > 1 else None
        if uni in anchor_to_key:                      # "University - City" → university
            return anchor_to_key[uni]
        for a, k in anchor_to_key.items():            # single-campus prefix
            if a.startswith(n + " ") or n.startswith(a + " "):
                return k
        if city and city in city_to_key:              # city fallback
            return city_to_key[city]
        return None

    ranks: dict[int, int] = {}
    unmatched: list[str] = []
    for row in re.findall(r"<tr>(.*?)</tr>", section, re.S):
        mr = re.search(r'rank-forward">(\d+)<', row)
        mn = re.search(r'<td class="left">(.*?)</td>', row, re.S)
        if not (mr and mn):
            continue
        name = re.sub(r"<[^>]+>", "", mn.group(1)).replace("&amp;", "&").strip()
        k = resolve(name)
        if k is None:
            unmatched.append(name)
        elif k not in ranks:                          # first (Values tab) wins
            ranks[k] = int(mr.group(1))
    print(f"  forward_ranks: {len(ranks)} markets ranked from forward-model.html")
    if unmatched:
        print(f"    {len(unmatched)} screener markets not on the dashboard: "
              + "; ".join(unmatched))
    return ranks


def connect(auth: str = "aad"):
    drivers = sorted(
        [d for d in pyodbc.drivers() if d.startswith("ODBC Driver")],
        reverse=True,
    )
    if not drivers:
        sys.exit("ODBC Driver 17 or 18 for SQL Server is not installed.")
    base = (
        f"Driver={{{drivers[0]}}};"
        f"Server=tcp:{SERVER},1433;Database={DATABASE};"
        "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30;"
    )
    if auth == "integrated":
        # Uses the Windows-logged-in user's AAD token silently. Works on
        # Entra-joined boxes; the script can run unattended.
        cs = base + "Authentication=ActiveDirectoryIntegrated;"
    elif auth == "aad":
        upn = input("Your @subtextliving.com email (Azure AD UPN): ").strip()
        if not upn:
            sys.exit("No UPN provided; aborting.")
        cs = base + f"UID={upn};Authentication=ActiveDirectoryInteractive;"
    elif auth == "sql":
        uid = input("SQL username: ").strip()
        pwd = getpass.getpass("SQL password (not echoed): ")
        if not uid or not pwd:
            sys.exit("Empty username or password; aborting.")
        cs = base + f"UID={uid};PWD={pwd};"
    elif auth == "env":
        uid = os.environ.get("SQLUSER")
        pwd = os.environ.get("SQLPASSWORD")
        if not uid or not pwd:
            sys.exit("--auth env requires SQLUSER and SQLPASSWORD env vars to be set.")
        cs = base + f"UID={uid};PWD={pwd};"
    else:
        sys.exit(f"Unknown --auth mode: {auth}")
    return pyodbc.connect(cs, timeout=30)


EXCEL_PATH = Path(
    r"C:\Users\JakeMillenbruck\Subtext\Subtext - Documents"
    r"\General\Markets\Support\Market Analysis Schedule.xlsx"
)

# Map the Schedule sheet's row-2 headers to clean JSON keys.
# Two columns share the label "Decision" - we disambiguate by position.
SCHEDULE_COLUMNS = {
    2: "market_type",            # col B
    3: "market_name",            # col C ("New Market Review")
    4: "analyst",                # col D ("Analyst/Intern")
    5: "initial_analysis_date",  # col E
    8: "initial_decision",       # col H ("Decision" #1)
    9: "ic_date",                # col I
    10: "ic_decision",           # col J ("Decision" #2)
    11: "status",                # col K
    12: "est_sites",             # col L
    13: "notes",                 # col M
}


def _prior_table(name: str) -> list[dict]:
    """Return a table from the existing data.json, or [] if unavailable.

    Used as a graceful fallback when a fresh read can't complete - keeps the
    last-known data instead of blanking the table on the live site."""
    try:
        if OUTPUT.exists():
            prev = json.loads(OUTPUT.read_text(encoding="utf-8"))
            return prev.get("tables", {}).get(name, [])
    except Exception as e:  # noqa: BLE001 - fallback must never raise
        print(f"  (could not read prior {name} from data.json: {e})")
    return []


def _call_with_timeout(fn, timeout_s: float, fallback):
    """Run fn() in a daemon thread; if it doesn't finish within timeout_s (or
    raises), log and return fallback(). Guards against blocking I/O - notably a
    OneDrive Files-On-Demand placeholder stalling indefinitely on cloud recall,
    which previously hung the unattended weekly refresh until Task Scheduler
    killed it at its 30-minute limit."""
    box: dict = {}

    def worker():
        try:
            box["value"] = fn()
        except Exception as e:  # noqa: BLE001
            box["error"] = e

    t = threading.Thread(target=worker, name=getattr(fn, "__name__", "worker"), daemon=True)
    t.start()
    t.join(timeout_s)
    if t.is_alive():
        print(f"  WARNING: {t.name} did not finish within {timeout_s:.0f}s "
              f"(likely a OneDrive cloud-recall stall) - using last-known data")
        return fallback()
    if "error" in box:
        err = box["error"]
        print(f"  WARNING: {t.name} failed ({type(err).__name__}: {err}) "
              f"- using last-known data")
        return fallback()
    return box["value"]


def read_market_analysis_excel() -> list[dict]:
    """Read the Market Analysis Schedule sheet into a list of row dicts.

    Column A on the sheet contains section headers (e.g., 'Deferred and Assessing
    Markets', 'Approved Markets'). Data rows follow each section header with
    column A blank. We attach the most-recent section header as `category` on
    every data row that follows.

    Missing file is non-fatal - returns []."""
    if not EXCEL_PATH.exists():
        print(f"  Excel not found at {EXCEL_PATH} (skipping)")
        return []
    try:
        import openpyxl  # local import - only needed if file is present
    except ImportError:
        print("  openpyxl not installed (skipping Excel)")
        return []

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    if "Market Analysis Schedule" not in wb.sheetnames:
        print("  'Market Analysis Schedule' sheet missing (skipping)")
        return []
    ws = wb["Market Analysis Schedule"]

    rows_out: list[dict] = []
    current_category: str | None = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= 2:
            continue  # skip blank row 1 and header row 2

        col_a = row[0]
        data_cols = [row[idx - 1] for idx in SCHEDULE_COLUMNS]

        # Section header: col A populated, other tracked cols empty
        if col_a and not any(v not in (None, "") for v in data_cols):
            current_category = str(col_a).strip()
            continue

        # Skip rows that are entirely empty across tracked cols
        if not any(v not in (None, "") for v in data_cols):
            continue

        # Skip per-section header rows that repeat the column labels.
        # These show up at the start of each section in the sheet with values
        # like market_type='Market Type', market_name='Market', etc.
        mt = (row[1] or "").strip() if isinstance(row[1], str) else row[1]
        if mt in ("Market Type",):
            continue

        row_dict: dict = {"category": current_category}
        for col_idx, key in SCHEDULE_COLUMNS.items():
            v = row[col_idx - 1]
            if isinstance(v, dt.datetime):
                row_dict[key] = v.date().isoformat()
            elif isinstance(v, dt.date):
                row_dict[key] = v.isoformat()
            elif v is None:
                row_dict[key] = None
            else:
                row_dict[key] = str(v).strip() if isinstance(v, str) else v
        rows_out.append(row_dict)

    print(f"  market_analysis_schedule  {len(rows_out)} rows  ({EXCEL_PATH.name})")
    return rows_out


# ============================================================
# Subtext Qualifier Engine
# ------------------------------------------------------------
# Criteria come from the "Market Analysis Qualifiers" sheet of
# Market Analysis Schedule.xlsx. Each evaluator computes a market's
# actual value and compares to the threshold. Returns:
#   {"id", "label", "threshold_display", "actual_display", "actual",
#    "status": "pass" | "fail" | "na", "explanation"}
# Phase-2 evaluators return status="na" until we wire their data source.
# ============================================================

def _na(qid, label, threshold_display, reason="Phase 2 - data not yet loaded"):
    return {
        "id": qid, "label": label,
        "threshold_display": threshold_display,
        "actual_display": "-", "actual": None,
        "status": "na", "tier": "na", "explanation": reason,
    }


def _tier(actual, threshold, margin, direction):
    """Binary tier: returns the same pass/fail as status. The historical
    'warn' middle band was retired - the scorecard now treats each
    qualifier as a clean pass or fail."""
    if direction == "above":
        return "pass" if actual > threshold else "fail"
    if direction == "below":
        return "pass" if actual <= threshold else "fail"
    return "fail"


def _result(qid, label, threshold_display, actual, actual_display,
            threshold, margin, direction, explanation):
    """Compute pass/fail status. `margin` is accepted for backwards
    compatibility with old call sites but is no longer used."""
    if direction == "above":
        passed = actual > threshold
    else:  # "below"
        passed = actual <= threshold
    status = "pass" if passed else "fail"
    return {
        "id": qid, "label": label,
        "threshold_display": threshold_display,
        "actual_display": actual_display, "actual": actual,
        "status": status,
        "tier": status,
        "explanation": explanation,
    }


def _q_rent_market(market, _props_by_market):
    rent = market.get("avg_rent_per_bed")
    if rent is None:
        return _na("rent_market", "Nominal market rent above $900", "> $900",
                   "no market rent on file")
    rent = float(rent)
    return _result(
        "rent_market", "Nominal market rent above $900", "> $900",
        actual=rent, actual_display=f"${rent:,.0f}",
        threshold=900, margin=90, direction="above",
        explanation="from MarketReports.rate_avg",
    )


def _q_rent_compset(market, props_by_market):
    # Comp-set is computed centrally before evaluators run (see main()) and
    # each market dict carries `comp_set_keys`. Filter the property list down
    # to that set, then average rent.
    mk = market["market_key"]
    comp_keys = market.get("comp_set_keys") or set()
    rents = [
        p["avg_rent"]
        for p in props_by_market.get(mk, [])
        if p.get("property_key") in comp_keys and p.get("avg_rent")
    ]
    if not rents:
        return _na("rent_compset", "Comp-set rent above $1,000", "> $1,000",
                   "no comp-set properties with rent")
    avg = sum(rents) / len(rents)
    return _result(
        "rent_compset", "Comp-set rent above $1,000", "> $1,000",
        actual=avg, actual_display=f"${avg:,.0f}",
        threshold=1000, margin=100, direction="above",
        explanation=f"average across {len(rents)} comp-set properties",
    )


def _q_fte(market, _props_by_market):
    fte = market.get("enr_full_time")
    if fte is None:
        return _na("fte", "FTE enrollment above 15,000", "> 15,000",
                   "no FTE on file")
    return _result(
        "fte", "FTE enrollment above 15,000", "> 15,000",
        actual=fte, actual_display=f"{int(fte):,}",
        threshold=15000, margin=1500, direction="above",
        explanation="current-year FTE from Enrollments_Manual",
    )


def _q_occupancy(market, _props_by_market):
    occ = market.get("occupancy")
    if occ is None:
        return _na("occupancy", "Market occupancy above 90%", "> 90%",
                   "no occupancy on file")
    occ = float(occ)
    return _result(
        "occupancy", "Market occupancy above 90%", "> 90%",
        actual=occ, actual_display=f"{occ * 100:.1f}%",
        threshold=0.90, margin=0.02, direction="above",
        explanation="from MarketReports.occupancy",
    )


def _q_pipeline(market, _props_by_market):
    pipe = market.get("beds_pipeline_total")
    existing = market.get("existing_beds")
    if not existing or pipe is None:
        return _na("pipeline", "Pipeline ≤ 10% of existing supply", "≤ 10% of existing",
                   "no supply data on file")
    ratio = pipe / existing
    return _result(
        "pipeline", "Pipeline ≤ 10% of existing supply", "≤ 10% of existing",
        actual=ratio, actual_display=f"{ratio * 100:.1f}%",
        threshold=0.10, margin=0.02, direction="below",
        explanation=f"{pipe:,} pipeline / {existing:,} existing PBSH beds",
    )


# --- Phase 2 placeholders (need additional data pulls) -----------------

def _q_fte_growth_2022(market, _props_by_market):
    fte_hist = market.get("fte_history")
    if not fte_hist or fte_hist.get("fte_growth_since_2022") is None:
        return _na("fte_growth_2022", "FTE growth since 2022 above 3%", "≥ 3% since 2022",
                   "no 2021-22 FTE baseline")
    g = float(fte_hist["fte_growth_since_2022"])
    sign = "+" if g >= 0 else ""
    baseline_yr = (fte_hist.get("baseline_2022_snapshot") or "")[:4]
    return _result(
        "fte_growth_2022", "FTE growth since 2022 above 3%", "≥ 3% since 2022",
        actual=g, actual_display=f"{sign}{g * 100:.1f}%",
        threshold=0.03, margin=0.005, direction="above",
        explanation=(
            f"current FTE {int(fte_hist['current_fte']):,} vs "
            f"{int(fte_hist['baseline_2022_fte']):,} ({baseline_yr})"
        ),
    )


def _q_fte_growth_yoy(market, _props_by_market):
    fte_hist = market.get("fte_history")
    if not fte_hist or fte_hist.get("yoy_fte_growth") is None:
        return _na("fte_growth_yoy", "FTE growth YoY positive", "> 0% YoY",
                   "no prior FTE snapshot with a different value")
    yoy = float(fte_hist["yoy_fte_growth"])
    sign = "+" if yoy >= 0 else ""
    prior_snap = (fte_hist.get("prior_year_snapshot") or "")[:10]
    return _result(
        "fte_growth_yoy", "FTE growth YoY positive", "> 0% YoY",
        actual=yoy, actual_display=f"{sign}{yoy * 100:.1f}%",
        threshold=0.0, margin=0.005, direction="above",
        explanation=(
            f"current FTE {int(fte_hist['current_fte']):,} vs "
            f"{int(fte_hist['prior_year_fte']):,} as of {prior_snap} "
            "(most recent snapshot with a different value)"
        ),
    )

def _q_rent_growth_3yr(market, _props_by_market):
    # Pull all available years of avg_rent_per_bed from market_history, take
    # the latest four to derive three trailing YoY rates, and require each to
    # be ≥ 3%.
    hist = market.get("market_history_rows") or []
    by_year = {int(r["year_"]): r.get("avg_rent_per_bed") for r in hist
               if r.get("avg_rent_per_bed") is not None
               and float(r["avg_rent_per_bed"]) > 0}
    if len(by_year) < 4:
        return _na("rent_growth_3yr", "YoY rent growth ≥3% for trailing 3 years",
                   "≥ 3% for 3 trailing years",
                   f"only {len(by_year)} years of non-zero rent history (need 4)")
    years = sorted(by_year.keys())[-4:]
    rents = [float(by_year[y]) for y in years]
    yoys = [(rents[i] - rents[i - 1]) / rents[i - 1] for i in (1, 2, 3)]
    worst = min(yoys)
    breakdown = [
        {
            "label": f"{years[i - 1]}-{years[i]}",
            "year_from": years[i - 1],
            "year_to": years[i],
            "yoy": y,
            "yoy_display": f"{y * 100:+.1f}%",
            "passed": bool(y >= 0.03),
        }
        for i, y in zip((1, 2, 3), yoys)
    ]
    detail = " · ".join(f"{b['year_from']}→{b['year_to']}: {b['yoy_display']}" for b in breakdown)
    res = _result(
        "rent_growth_3yr", "YoY rent growth ≥3% for trailing 3 years",
        "≥ 3% for 3 trailing years",
        actual=worst,
        actual_display=" · ".join(b["yoy_display"] for b in breakdown),
        threshold=0.03, margin=0.005, direction="above",
        explanation=detail,
    )
    res["breakdown"] = breakdown
    return res

def _q_prelease_vs_prior(market, _props_by_market):
    yoy = market.get("prelease_yoy")
    if not yoy:
        return _na("prelease_lag", "Market prelease not lagging prior year by >5%", "delta ≥ −5%",
                   "no same-week prior-cycle observation")
    delta = yoy["delta"]
    sign = "+" if delta >= 0 else ""
    cur_pct = yoy["current_prelease"] * 100
    prior_pct = yoy["prior_prelease"] * 100
    return _result(
        "prelease_lag", "Market prelease not lagging prior year by >5%", "delta ≥ −5%",
        actual=delta, actual_display=f"{sign}{delta * 100:.1f}%",
        threshold=-0.05, margin=0.01, direction="above",
        explanation=(
            f"week {yoy['current_week']} of cycle {yoy['current_cycle']}: "
            f"{cur_pct:.1f}% vs {prior_pct:.1f}% same week in "
            f"cycle {yoy['current_cycle'] - 1}"
        ),
    )

def _q_uncaptured_demand(market, props_by_market):
    # Approximate uncaptured demand near campus as the share of FTE not yet
    # served by purpose-built supply within 1 mile of campus.
    #   beds_1mi   = sum of beds in properties with milesToClosestCampus ≤ 1
    #   captured   = beds_1mi / FTE
    #   uncaptured = 1 - captured
    # Markets with high uncaptured demand still have room for new PBSH supply.
    fte = market.get("enr_full_time")
    if not fte:
        return _na("uncaptured_1mi", "Uncaptured demand within 1 mile above 30%",
                   "> 30%", "no FTE on file")
    mk = market["market_key"]
    near = [p for p in props_by_market.get(mk, [])
            if p.get("milesToClosestCampus") is not None
            and float(p["milesToClosestCampus"]) <= 1.0
            and p.get("beds")]
    beds_1mi = sum(int(p["beds"]) for p in near)
    captured = beds_1mi / float(fte)
    uncaptured = 1.0 - captured
    return _result(
        "uncaptured_1mi", "Uncaptured demand within 1 mile above 30%", "> 30%",
        actual=uncaptured, actual_display=f"{uncaptured * 100:.1f}%",
        threshold=0.30, margin=0.02, direction="above",
        explanation=(
            f"{beds_1mi:,} PBSH beds within 1 mi of campus "
            f"({len(near)} props) vs {int(fte):,} FTE - "
            f"capture rate {captured * 100:.1f}%"
        ),
    )

def _q_income(market, _props_by_market):
    inc = market.get("mean_origin_income")
    n = market.get("affluence_n_students")
    if inc is None or (n is not None and n < 100):
        return _na("income_median_zips", "Avg income of median zipcodes above $85K", "> $85,000",
                   "no migration sample" if inc is None else f"sample too small (n={n})")
    return _result(
        "income_median_zips", "Avg income of median zipcodes above $85K", "> $85,000",
        actual=inc, actual_display=f"${inc:,.0f}",
        threshold=85000, margin=5000, direction="above",
        explanation=(
            f"mean origin household income across {int(n):,} students "
            "(migration data, 2023-07 + 2023-08)"
        ) if n is not None else "mean origin household income (migration data)",
    )


# --- Power 4 / R1 + forward-ranking qualifiers ------------------------
# Anchor classification and the qualifier-result shape live in
# university_classification.py so the standalone data.json patcher
# (load_university_qualifiers.py) produces identical results.

_POWER4_CACHE = None


def _power4_anchors_cached():
    global _POWER4_CACHE
    if _POWER4_CACHE is None:
        from university_classification import load_power4_anchors
        _POWER4_CACHE = load_power4_anchors()
    return _POWER4_CACHE


def _q_power4_r1(market, _props_by_market):
    from university_classification import power4_r1_result
    return power4_r1_result(market.get("anchor_university"), _power4_anchors_cached())


def _q_fwd_top50(market, _props_by_market):
    # fwd_rank is stamped onto the scorecard from forward-model.html earlier in
    # main(), before compute_qualifiers runs.
    from university_classification import forward_top50_result
    return forward_top50_result(market.get("fwd_rank"))


# Display order matches the Excel sheet.
QUALIFIER_EVALUATORS = [
    _q_rent_market,
    _q_rent_compset,
    _q_fte,
    _q_occupancy,
    _q_fte_growth_2022,
    _q_fte_growth_yoy,
    _q_rent_growth_3yr,
    _q_prelease_vs_prior,
    _q_pipeline,
    _q_uncaptured_demand,
    _q_income,
    _q_power4_r1,
    _q_fwd_top50,
]


def compute_qualifiers(tables: dict) -> list[dict]:
    """Return list of {market_key, results, passes, evaluable, total, score_pct}."""
    sc = tables.get("scorecard", [])
    props = tables.get("properties", [])

    by_market: dict = {}
    for p in props:
        by_market.setdefault(p["market_key"], []).append(p)

    # Affluence is computed earlier in main(); merge those fields into each
    # market dict so income-driven evaluators can read them.
    affluence_by_market = {
        a["market_key"]: a for a in tables.get("market_affluence", [])
    }
    # Prelease YoY is derived from the prelease_velocity table.
    from load_affluence import compute_prelease_yoy
    prelease_yoy_by_market = compute_prelease_yoy(tables.get("prelease_velocity", []))

    # FTE history from the new fte_history query.
    fte_history_by_market = {
        r["market_key"]: r for r in tables.get("fte_history", [])
    }
    # Multi-year rent + supply snapshots, grouped by market for the
    # rent_growth_3yr qualifier.
    market_history_by_market: dict = {}
    for r in tables.get("market_history", []):
        market_history_by_market.setdefault(r["market_key"], []).append(r)

    # Comp-set membership per market (see load_comp_set.py for definition).
    from load_comp_set import build_comp_set_keys
    comp_keys_by_market = build_comp_set_keys(by_market)
    # Also stamp is_comp_set onto each property row so the UI / table can
    # show comp-set membership.
    for p in props:
        p["is_comp_set"] = p.get("property_key") in comp_keys_by_market.get(p["market_key"], set())

    out: list[dict] = []
    for market in sc:
        m = dict(market)
        a = affluence_by_market.get(m["market_key"])
        if a is not None:
            m["mean_origin_income"] = a.get("mean_origin_income")
            m["affluence_n_students"] = a.get("n_students")
        m["prelease_yoy"] = prelease_yoy_by_market.get(m["market_key"])
        m["fte_history"] = fte_history_by_market.get(m["market_key"])
        m["market_history_rows"] = market_history_by_market.get(m["market_key"], [])
        m["comp_set_keys"] = comp_keys_by_market.get(m["market_key"], set())
        results = [ev(m, by_market) for ev in QUALIFIER_EVALUATORS]
        q = {
            "market_key": m["market_key"],
            "results": results,
            "total": len(results),
        }
        # Weighted rollup - qualifiers with a `breakdown` list (e.g.
        # rent_growth_3yr) award fractional credit. See load_affluence.py.
        from load_affluence import recompute_rollup
        recompute_rollup(q)
        out.append(q)
    placeholders = sum(1 for r in out[0]["results"] if r["status"] == "na") if out else 0
    print(f"  market_qualifiers computed for {len(out)} markets "
          f"({len(QUALIFIER_EVALUATORS) - placeholders} live, {placeholders} placeholders)")
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Export Azure SQL data to data.json.")
    parser.add_argument(
        "--auth",
        choices=["aad", "integrated", "sql", "env"],
        default="aad",
        help="Auth mode (default: aad - Azure AD interactive). "
             "'integrated' uses Windows AAD silently; "
             "'sql' prompts via getpass; 'env' reads SQLUSER/SQLPASSWORD.",
    )
    args = parser.parse_args()

    cn = connect(args.auth)
    cur = cn.cursor()

    payload: dict = {
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source": f"{DATABASE} @ {SERVER}",
        "tables": {},
    }

    for name, sql in QUERIES.items():
        print(f"  {name} ...", end="", flush=True)
        cur.execute(sql)
        cols = [c[0] for c in cur.description]
        rows = [
            {col: to_jsonable(val) for col, val in zip(cols, row)}
            for row in cur.fetchall()
        ]
        payload["tables"][name] = rows
        print(f" {len(rows)} rows")

    # --- Flag active markets (CRM 'Markets - *' board: Upcoming/Assessing/Pursuing) ---
    # Stamp market_status onto the scorecard so the Industry-page "Active markets"
    # toggle can filter + colour the map off a single field. is_pursuit is kept
    # (derived from status) for backward compat with the market-state report.
    # NOTE: pinned from ACTIVE_MARKET_STATUS for now, not SQL - see that constant.
    fwd_ranks = compute_forward_ranks(payload["tables"]["scorecard"])
    for r in payload["tables"]["scorecard"]:
        status = ACTIVE_MARKET_STATUS.get(r["market_key"])
        r["market_status"] = status               # "pursuing"|"assessing"|"upcoming"|None
        r["is_pursuit"] = 1 if status == "pursuing" else 0
        r["pursuit_deals"] = 1 if status == "pursuing" else 0
        r["fwd_rank"] = fwd_ranks.get(r["market_key"])  # None if not in the screener
    _sc = payload["tables"]["scorecard"]
    print("  active markets: "
          f"{sum(1 for r in _sc if r.get('market_status') == 'upcoming')} upcoming, "
          f"{sum(1 for r in _sc if r.get('market_status') == 'assessing')} assessing, "
          f"{sum(1 for r in _sc if r.get('market_status') == 'pursuing')} pursuing")

    # --- Read the Market Analysis Schedule Excel (OneDrive synced file) ---
    # Guarded with a timeout: the file is a OneDrive Files-On-Demand placeholder,
    # and an unattended run can stall indefinitely on its cloud recall. On
    # timeout/error we keep the last-known schedule from the prior data.json so
    # one stalled read can't blank the table or hang the whole refresh.
    payload["tables"]["market_analysis_schedule"] = _call_with_timeout(
        read_market_analysis_excel,
        120,
        lambda: _prior_table("market_analysis_schedule"),
    )

    # --- Compute per-market affluence from migration CSV ---
    # Source: ../Affluence Data/MigrationAllRents*.csv (newest file).
    # Logic lives in load_affluence.py so it can also run standalone
    # without an Azure SQL refresh. Must run BEFORE compute_qualifiers so
    # the income qualifier can read mean_origin_income.
    from load_affluence import (
        build_market_affluence, newest_csv, DEFAULT_CSV_DIR, DEFAULT_MONTHS,
    )
    try:
        aff_csv = newest_csv(DEFAULT_CSV_DIR)
        ipeds_to_market = {
            int(r["ipeds_id"]): int(r["market_key"])
            for r in payload["tables"].get("campus_locations", [])
            if r.get("ipeds_id") is not None and r.get("market_key") is not None
        }
        payload["tables"]["market_affluence"] = build_market_affluence(
            aff_csv, ipeds_to_market, months=list(DEFAULT_MONTHS),
        )
        print(f"  market_affluence from {aff_csv.name} "
              f"[months={','.join(DEFAULT_MONTHS)}]: "
              f"{len(payload['tables']['market_affluence'])} markets")
    except SystemExit as e:
        print(f"  market_affluence SKIPPED: {e}")
        payload["tables"]["market_affluence"] = []

    # --- Compute Subtext qualifier scorecard per market ---
    payload["tables"]["market_qualifiers"] = compute_qualifiers(payload["tables"])

    # --- Fetch this week's Subtext Dispatch headlines (for the ticker) ---
    # Lives at the top level (not under 'tables') because it's a single
    # object, not a row-set. Non-fatal: dispatch outages must not block
    # the weekly SQL refresh.
    try:
        from load_dispatch import fetch_dispatch_headlines
        payload["dispatch_headlines"] = fetch_dispatch_headlines()
        print(f"  dispatch_headlines: "
              f"{len(payload['dispatch_headlines']['features'])} features + "
              f"{len(payload['dispatch_headlines']['briefs'])} briefs "
              f"[{payload['dispatch_headlines']['issue']}]")
    except Exception as e:
        print(f"  dispatch_headlines SKIPPED: {type(e).__name__}: {e}")
        payload["dispatch_headlines"] = None

    # --- Split: plans → per-property JSON files (loaded on demand) ---
    plans_dir = OUTPUT.parent / "plans"
    plans_dir.mkdir(exist_ok=True)
    plans = payload["tables"].pop("plans", [])
    by_property: dict[int, list[dict]] = {}
    for row in plans:
        pk = row.get("property_key")
        if pk is None:
            continue
        by_property.setdefault(pk, []).append(row)
    # Clean stale files before writing the new set
    for old in plans_dir.glob("*.json"):
        old.unlink()
    for pk, rows in by_property.items():
        (plans_dir / f"{pk}.json").write_text(
            json.dumps(rows, default=str), encoding="utf-8",
        )
    print(f"  plans split into {len(by_property)} per-property files in {plans_dir}/")

    # --- unit_mix: per-property beds/units by bedroom type --------------
    # Built from the same in-memory plan rows (single source of truth lives
    # in load_unit_mix.py). Drives the Unit & Bed Mix section on market.html.
    from load_unit_mix import mix_for_property
    unit_mix = []
    for pk, rows in by_property.items():
        mix = mix_for_property(rows)
        if mix["total_beds"] == 0 and mix["total_units"] == 0:
            continue
        market_key = next((r.get("market_key") for r in rows if r.get("market_key") is not None), None)
        unit_mix.append({"property_key": pk, "market_key": market_key, **mix})
    unit_mix.sort(key=lambda r: (r["market_key"] or 0, -r["total_beds"]))
    payload["tables"]["unit_mix"] = unit_mix
    print(f"  unit_mix: {len(unit_mix)} properties with bedroom-type mix")

    # Derived dashboard-level data_as_of = min across all tables that carry one
    as_ofs = []
    for table in payload["tables"].values():
        for row in table:
            for k, v in row.items():
                if k in ("data_as_of", "current_snapshot") and v:
                    as_ofs.append(v)
    payload["data_as_of"] = max(as_ofs) if as_ofs else None

    # --- Manual overrides (overrides.json) ---------------------------------
    # Applied LAST so hand corrections for known-bad source figures survive
    # every refresh. Managed via apply_overrides.py; never edit data.json
    # directly. A failure here must not sink the whole refresh.
    try:
        from apply_overrides import load_overrides, apply_overrides
        _overrides = load_overrides()
        if _overrides:
            print(f"\nApplying {len(_overrides)} manual override(s) from overrides.json:")
            for _line in apply_overrides(payload, _overrides):
                print(f"  {_line}")
    except Exception as e:
        print(f"  overrides SKIPPED: {type(e).__name__}: {e}")

    with OUTPUT.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, default=str)

    size_kb = OUTPUT.stat().st_size / 1024
    print(f"\nWrote {OUTPUT} ({size_kb:.0f} KB)")
    print(f"data_as_of = {payload['data_as_of']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
