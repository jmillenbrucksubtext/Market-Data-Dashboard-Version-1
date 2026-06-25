"""
Patch tables.admissions_history into data.json from the Excel repository.

Reads "Admissions Data/admissions_history.xlsx" (built by
build_admissions_repository.py from the IPEDS ADM survey) and writes one row
per (ipeds_id, year) into data.json. The University tab's "Applications,
Acceptance & Yield" chart joins it by ipeds_id and computes acceptance rate
(admitted / applications) and yield (enrolled / admitted) client-side.

Slim data.json patcher in the same spirit as load_affluence.py - no SQL. Run
after export-data.py so the weekly refresh doesn't wipe the table.

Run:
    python load_admissions_history.py
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
DATA_JSON = HERE / "data.json"
XLSX = HERE / "Admissions Data" / "admissions_history.xlsx"


def build_rows() -> list[dict]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    idx = {name: i for i, name in enumerate(header)}

    def cell(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    def as_int(v):
        try:
            return int(round(float(v)))
        except (ValueError, TypeError):
            return None

    rows: list[dict] = []
    for row in rows_iter:
        if row is None or cell(row, "ipeds_id") is None:
            continue
        ip = as_int(cell(row, "ipeds_id"))
        yr = as_int(cell(row, "year"))
        if ip is None or yr is None:
            continue
        appl = as_int(cell(row, "applications"))
        adm = as_int(cell(row, "admitted"))
        enr = as_int(cell(row, "enrolled"))
        if appl is None and adm is None and enr is None:
            continue
        rows.append({
            "ipeds_id": ip,
            "year_": yr,
            "applications": appl,
            "admitted": adm,
            "enrolled": enr,
        })
    rows.sort(key=lambda r: (r["ipeds_id"], r["year_"]))
    return rows


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX} - run build_admissions_repository.py first.")
        return 1
    payload = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    rows = build_rows()
    payload["tables"]["admissions_history"] = rows

    schools = len({r["ipeds_id"] for r in rows})
    years = sorted({r["year_"] for r in rows})
    DATA_JSON.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    print(
        f"admissions_history: {len(rows)} rows · {schools} schools"
        f" · years {years[0]}-{years[-1]}" if rows else "admissions_history: 0 rows",
        f"Wrote {DATA_JSON} ({DATA_JSON.stat().st_size / 1024:.0f} KB)",
        sep="\n",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
