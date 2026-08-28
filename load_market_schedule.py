"""
load_market_schedule.py
-----------------------
Re-reads Market Analysis Schedule.xlsx (the team's live market-analysis
tracker on OneDrive) and patches tables.market_analysis_schedule in
data.json, feeding the Analysis Schedule tab on the Industry page.

The reader mirrors read_market_analysis_excel() in export-data.py, which
does the same job during the weekly refresh - run this standalone when the
workbook has changed and you don't want to wait for Monday (matches the
load_*.py qualifier-patcher pattern).

Usage (standalone):
    python load_market_schedule.py
"""

from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

DATA_JSON = Path(__file__).parent / "data.json"

EXCEL_PATH = Path(
    r"C:\Users\JakeMillenbruck\Subtext\Subtext - Documents"
    r"\General\Markets\Support\Market Analysis Schedule.xlsx"
)

# Map the Schedule sheet's row-2 headers to clean JSON keys.
# Two columns share the label "Decision" - disambiguated by position.
# Keep in sync with SCHEDULE_COLUMNS in export-data.py.
SCHEDULE_COLUMNS = {
    2: "market_type",            # col B
    3: "market_name",            # col C ("New Market Review")
    4: "analyst",                # col D ("Analyst/Intern")
    5: "initial_analysis_date",  # col E
    8: "initial_decision",       # col H ("Decision" #1)
    9: "ic_date",                # col I
    10: "ic_decision",           # col J ("Decision" #2)
    11: "status",                # col K
    12: "est_sites",             # col L
    13: "notes",                 # col M
}


def read_market_analysis_excel(path: Path = EXCEL_PATH) -> list[dict]:
    """Read the Market Analysis Schedule sheet into a list of row dicts.

    Column A holds section headers ('Deferred and Assessing Markets', ...);
    data rows follow with column A blank and get the most-recent header
    attached as `category`. Raises on a missing file - standalone runs
    should fail loudly rather than silently blank the tab."""
    import openpyxl

    if not path.exists():
        sys.exit(f"workbook not found at {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Market Analysis Schedule" not in wb.sheetnames:
        sys.exit("'Market Analysis Schedule' sheet missing from workbook")
    ws = wb["Market Analysis Schedule"]

    rows_out: list[dict] = []
    current_category: str | None = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= 2:
            continue  # skip blank row 1 and header row 2

        col_a = row[0]
        data_cols = [row[idx - 1] for idx in SCHEDULE_COLUMNS]

        # Section header: col A populated, other tracked cols empty
        if col_a and not any(v not in (None, "") for v in data_cols):
            current_category = str(col_a).strip()
            continue

        # Skip rows that are entirely empty across tracked cols
        if not any(v not in (None, "") for v in data_cols):
            continue

        # Skip per-section header rows that repeat the column labels
        mt = (row[1] or "").strip() if isinstance(row[1], str) else row[1]
        if mt in ("Market Type",):
            continue

        row_dict: dict = {"category": current_category}
        for col_idx, key in SCHEDULE_COLUMNS.items():
            v = row[col_idx - 1]
            if isinstance(v, dt.datetime):
                row_dict[key] = v.date().isoformat()
            elif isinstance(v, dt.date):
                row_dict[key] = v.isoformat()
            elif v is None:
                row_dict[key] = None
            else:
                row_dict[key] = str(v).strip() if isinstance(v, str) else v
        rows_out.append(row_dict)

    return rows_out


def patch_data_json(data_path: Path = DATA_JSON) -> list[dict]:
    """Standalone-run entry: refresh tables.market_analysis_schedule."""
    if not data_path.exists():
        sys.exit(f"data.json not found at {data_path}")
    rows = read_market_analysis_excel()
    payload = json.loads(data_path.read_text(encoding="utf-8"))
    payload.setdefault("tables", {})["market_analysis_schedule"] = rows
    data_path.write_text(
        json.dumps(payload, separators=(",", ":"), ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"  market_analysis_schedule: {len(rows)} rows  ({EXCEL_PATH.name})")
    return rows


if __name__ == "__main__":
    patch_data_json()
